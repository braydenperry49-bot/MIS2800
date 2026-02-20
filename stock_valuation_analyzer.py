#!/usr/bin/env python3
"""
Stock Valuation Analyzer — Enhanced Professional Edition
=========================================================
A comprehensive stock analysis tool that combines 7 valuation methods into a
weighted fair-value estimate, plus quality scoring, scenario analysis, and
risk assessment.  Exports a 20-sheet formatted Excel report.

Valuation Models:
    DCF (50%), Comparable Companies (40%), Historical P/E (10%)
    + Analyst Targets, Technical, Sentiment, Seasonal (informational)

Additional Analyses:
    Quality Score (0-100 across 6 dimensions), Scenario Analysis
    (bear/base/bull), Risk Assessment, Financial Statements, Peer Comparison

Usage:
    python stock_valuation_analyzer.py              # interactive prompt
    python stock_valuation_analyzer.py AAPL         # single ticker
    python stock_valuation_analyzer.py AAPL MSFT    # multiple tickers
"""

import argparse
import datetime
import hashlib
import json
import os
import sys
import time
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
import requests
import yfinance as yf
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import config as cfg

warnings.filterwarnings("ignore", category=FutureWarning)

# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — Caching Utilities
# ═══════════════════════════════════════════════════════════════════════════════

def _cache_path(key: str) -> str:
    os.makedirs(cfg.CACHE_DIR, exist_ok=True)
    hashed = hashlib.md5(key.encode()).hexdigest()
    return os.path.join(cfg.CACHE_DIR, f"{hashed}.json")


def cache_get(key: str):
    path = _cache_path(key)
    if not os.path.exists(path):
        return None
    try:
        with open(path, "r") as f:
            data = json.load(f)
        ts = data.get("_ts", 0)
        if (time.time() - ts) > cfg.CACHE_EXPIRY_HOURS * 3600:
            os.remove(path)
            return None
        return data.get("payload")
    except (json.JSONDecodeError, OSError):
        return None


def cache_set(key: str, payload):
    path = _cache_path(key)
    with open(path, "w") as f:
        json.dump({"_ts": time.time(), "payload": payload}, f)


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — Data Fetching Helpers
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_yfinance_data(ticker: str) -> dict:
    """Fetch comprehensive data from yfinance for a given ticker."""
    cache_key = f"yf_{ticker}"
    cached = cache_get(cache_key)
    if cached:
        return cached

    print(f"  [yfinance] Fetching data for {ticker}...")
    stock = yf.Ticker(ticker)

    info = {}
    try:
        info = stock.info or {}
    except Exception:
        pass

    # Historical prices — 5 years
    hist = pd.DataFrame()
    try:
        hist = stock.history(period="5y")
    except Exception:
        pass

    # Financial statements
    financials = {}
    for attr in ("income_stmt", "balance_sheet", "cashflow"):
        try:
            df = getattr(stock, attr, pd.DataFrame())
            if df is not None and not df.empty:
                financials[attr] = df.to_dict()
            else:
                financials[attr] = {}
        except Exception:
            financials[attr] = {}

    # Analyst recommendations
    recommendations = []
    try:
        rec = stock.recommendations
        if rec is not None and not rec.empty:
            recommendations = rec.tail(20).reset_index().to_dict(orient="records")
    except Exception:
        pass

    result = _serialize({
        "info": info,
        "history": hist.reset_index().to_dict(orient="list") if not hist.empty else {},
        "financials": financials,
        "recommendations": recommendations,
    })
    cache_set(cache_key, result)
    return result


def _serialize(obj):
    """Make an object JSON-serializable."""
    if isinstance(obj, dict):
        return {str(k): _serialize(v) for k, v in obj.items()}
    if isinstance(obj, (list, tuple)):
        return [_serialize(v) for v in obj]
    if isinstance(obj, (np.integer,)):
        return int(obj)
    if isinstance(obj, (np.floating, np.float64)):
        return float(obj)
    if isinstance(obj, (np.bool_,)):
        return bool(obj)
    if isinstance(obj, (np.ndarray,)):
        return _serialize(obj.tolist())
    if isinstance(obj, (pd.Timestamp, datetime.datetime, datetime.date)):
        return obj.isoformat()
    if isinstance(obj, pd.Timedelta):
        return str(obj)
    try:
        if pd.isna(obj):
            return None
    except (TypeError, ValueError):
        pass
    if isinstance(obj, (str, int, float, bool, type(None))):
        return obj
    return str(obj)


def fetch_fred_series(series_id: str):
    """Fetch latest value of a FRED series via the FRED API."""
    cache_key = f"fred_{series_id}"
    cached = cache_get(cache_key)
    if cached is not None:
        return cached

    api_key = cfg.FRED_API_KEY
    if not api_key or api_key == "demo":
        return None

    url = "https://api.stlouisfed.org/fred/series/observations"
    params = {
        "series_id": series_id,
        "api_key": api_key,
        "file_type": "json",
        "sort_order": "desc",
        "limit": 5,
    }
    try:
        resp = requests.get(url, params=params, timeout=10)
        resp.raise_for_status()
        observations = resp.json().get("observations", [])
        for obs in observations:
            val = obs.get("value", ".")
            if val != ".":
                result = float(val)
                cache_set(cache_key, result)
                return result
    except Exception:
        pass
    return None


def fetch_fmp_peers(ticker: str) -> list:
    """Get peer/comparable company tickers from Financial Modeling Prep."""
    cache_key = f"fmp_peers_{ticker}"
    cached = cache_get(cache_key)
    if cached is not None:
        return cached

    api_key = cfg.FMP_API_KEY
    if not api_key or api_key == "demo":
        return []

    url = f"https://financialmodelingprep.com/api/v4/stock_peers"
    params = {"symbol": ticker, "apikey": api_key}
    try:
        resp = requests.get(url, params=params, timeout=10)
        resp.raise_for_status()
        data = resp.json()
        if data and isinstance(data, list) and "peersList" in data[0]:
            peers = data[0]["peersList"][: cfg.COMPS_COUNT]
            cache_set(cache_key, peers)
            return peers
    except Exception:
        pass
    return []


def fetch_news_sentiment(ticker: str) -> list:
    """Fetch recent news headlines for sentiment analysis."""
    cache_key = f"news_{ticker}"
    cached = cache_get(cache_key)
    if cached is not None:
        return cached

    api_key = cfg.NEWS_API_KEY
    if not api_key:
        return []

    url = "https://newsapi.org/v2/everything"
    params = {
        "q": ticker,
        "language": "en",
        "sortBy": "publishedAt",
        "pageSize": 30,
        "apiKey": api_key,
    }
    try:
        resp = requests.get(url, params=params, timeout=10)
        resp.raise_for_status()
        articles = resp.json().get("articles", [])
        headlines = [
            {"title": a.get("title", ""), "description": a.get("description", "")}
            for a in articles
            if a.get("title")
        ]
        cache_set(cache_key, headlines)
        return headlines
    except Exception:
        return []


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 3 — Valuation Models
# ═══════════════════════════════════════════════════════════════════════════════

class ValuationResult:
    """Container for a single valuation method's output."""
    def __init__(self, method: str, fair_value: float, confidence: float,
                 details: dict = None):
        self.method = method
        self.fair_value = fair_value          # per-share value
        self.confidence = confidence          # 0-1 scale
        self.details = details or {}


# ── 3a. DCF Valuation ────────────────────────────────────────────────────────

def dcf_valuation(data: dict) -> ValuationResult:
    """Discounted Cash Flow valuation using free cash flow projections."""
    info = data.get("info", {})
    financials = data.get("financials", {})

    shares = info.get("sharesOutstanding") or info.get("impliedSharesOutstanding", 0)
    if not shares:
        return ValuationResult("DCF", 0, 0, {"error": "No shares outstanding data"})

    # Gather historical free cash flow
    cf_data = financials.get("cashflow", {})
    if not cf_data:
        return ValuationResult("DCF", 0, 0, {"error": "No cash flow data"})

    cf_df = pd.DataFrame(cf_data)
    fcf_row = None
    for label in ("Free Cash Flow", "FreeCashFlow"):
        if label in cf_df.index:
            fcf_row = cf_df.loc[label]
            break

    if fcf_row is None:
        # Try computing: Operating Cash Flow - CapEx
        op_cf = None
        capex = None
        for label in ("Operating Cash Flow", "OperatingCashFlow",
                       "Total Cash From Operating Activities"):
            if label in cf_df.index:
                op_cf = cf_df.loc[label]
                break
        for label in ("Capital Expenditure", "CapitalExpenditure",
                       "Capital Expenditures"):
            if label in cf_df.index:
                capex = cf_df.loc[label]
                break
        if op_cf is not None and capex is not None:
            fcf_row = op_cf.astype(float) + capex.astype(float)  # capex is negative
        else:
            return ValuationResult("DCF", 0, 0, {"error": "Cannot compute FCF"})

    fcf_values = [v for v in fcf_row.values if v is not None and not np.isnan(float(v))]
    if not fcf_values:
        return ValuationResult("DCF", 0, 0, {"error": "No valid FCF values"})

    latest_fcf = float(fcf_values[0])

    # Growth rate estimate
    if len(fcf_values) >= 2:
        growth_rates = []
        for i in range(len(fcf_values) - 1):
            prev = float(fcf_values[i + 1])
            curr = float(fcf_values[i])
            if prev > 0:
                growth_rates.append((curr - prev) / prev)
        avg_growth = np.mean(growth_rates) if growth_rates else 0.05
        avg_growth = np.clip(avg_growth, -0.05, 0.30)  # cap growth rate
    else:
        avg_growth = 0.05

    # Discount rate (WACC approximation)
    beta = info.get("beta", 1.0) or 1.0
    risk_free = cfg.DCF_DEFAULTS["risk_free_rate"]
    if risk_free is None:
        risk_free_fetched = fetch_fred_series("DGS10")
        risk_free = (risk_free_fetched / 100.0) if risk_free_fetched else 0.04
    erp = cfg.DCF_DEFAULTS["equity_risk_premium"]
    cost_of_equity = risk_free + beta * erp

    # Debt cost approximation
    total_debt = info.get("totalDebt", 0) or 0
    market_cap = info.get("marketCap", 0) or 0
    interest_expense = abs(info.get("interestExpense", 0) or 0)

    if total_debt > 0 and interest_expense > 0:
        cost_of_debt = interest_expense / total_debt
    else:
        cost_of_debt = risk_free + 0.02

    tax_rate = cfg.DCF_DEFAULTS["tax_rate"]
    total_value = market_cap + total_debt
    if total_value > 0:
        equity_weight = market_cap / total_value
        debt_weight = total_debt / total_value
    else:
        equity_weight, debt_weight = 1.0, 0.0

    wacc = equity_weight * cost_of_equity + debt_weight * cost_of_debt * (1 - tax_rate)
    wacc = max(wacc, 0.06)  # floor

    # Project FCFs
    proj_years = cfg.DCF_DEFAULTS["projection_years"]
    terminal_growth = cfg.DCF_DEFAULTS["terminal_growth_rate"]
    projected_fcfs = []
    for yr in range(1, proj_years + 1):
        projected = latest_fcf * ((1 + avg_growth) ** yr)
        projected_fcfs.append(projected)

    # Terminal value (Gordon Growth)
    terminal_fcf = projected_fcfs[-1] * (1 + terminal_growth)
    terminal_value = terminal_fcf / (wacc - terminal_growth) if wacc > terminal_growth else 0

    # Discount everything back
    pv_fcfs = sum(fcf / ((1 + wacc) ** yr) for yr, fcf in enumerate(projected_fcfs, 1))
    pv_terminal = terminal_value / ((1 + wacc) ** proj_years)
    enterprise_value = pv_fcfs + pv_terminal

    # Equity value
    cash = info.get("totalCash", 0) or 0
    equity_value = enterprise_value + cash - total_debt
    fair_value_per_share = equity_value / shares if shares > 0 else 0

    # Apply margin of safety
    mos = cfg.DCF_DEFAULTS["margin_of_safety"]
    buy_price = fair_value_per_share * (1 - mos)

    # Confidence based on data availability
    confidence = 0.7
    if len(fcf_values) >= 3:
        confidence += 0.1
    if total_debt > 0 and interest_expense > 0:
        confidence += 0.1
    confidence = min(confidence, 1.0)

    details = {
        "latest_fcf": latest_fcf,
        "fcf_growth_rate": avg_growth,
        "wacc": wacc,
        "cost_of_equity": cost_of_equity,
        "cost_of_debt": cost_of_debt,
        "beta": beta,
        "risk_free_rate": risk_free,
        "projected_fcfs": projected_fcfs,
        "terminal_value": terminal_value,
        "enterprise_value": enterprise_value,
        "equity_value": equity_value,
        "shares_outstanding": shares,
        "margin_of_safety": mos,
        "buy_price": buy_price,
    }

    return ValuationResult("DCF", fair_value_per_share, confidence, details)


# ── 3b. Comparable Companies ─────────────────────────────────────────────────

def comps_valuation(data: dict, ticker: str) -> ValuationResult:
    """Relative valuation using peer-company multiples."""
    info = data.get("info", {})
    peers = fetch_fmp_peers(ticker)

    # Fallback 1: use config-defined peer tickers
    if not peers:
        peers = cfg.PEER_TICKERS.get(ticker, [])

    # Fallback 2: use sector/industry from yfinance
    if not peers:
        industry_peers = info.get("recommendedSymbols", [])
        if industry_peers:
            peers = [p.get("symbol") for p in industry_peers if p.get("symbol")]

    if not peers:
        return ValuationResult("Comps", 0, 0, {"error": "No peer companies found"})

    peers = peers[: cfg.COMPS_COUNT]

    # Collect multiples for the target and peers
    target_multiples = _extract_multiples(info)
    peer_data = []

    for peer_ticker in peers:
        try:
            peer_stock = yf.Ticker(peer_ticker)
            peer_info = peer_stock.info or {}
            pm = _extract_multiples(peer_info)
            pm["ticker"] = peer_ticker
            pm["name"] = peer_info.get("shortName", peer_ticker)
            pm["marketCap"] = peer_info.get("marketCap", 0)
            peer_data.append(pm)
        except Exception:
            continue

    if not peer_data:
        return ValuationResult("Comps", 0, 0, {"error": "Could not fetch peer data"})

    # Calculate implied fair values from each multiple
    current_price = info.get("currentPrice") or info.get("regularMarketPrice", 0)
    implied_values = []
    multiple_details = {}

    for mult_name in cfg.COMPS_MULTIPLES:
        key = _multiple_key(mult_name)
        target_val = target_multiples.get(key)
        peer_vals = [p.get(key) for p in peer_data if p.get(key) and p[key] > 0]

        if not peer_vals or not target_val or target_val <= 0:
            continue

        median_peer = np.median(peer_vals)

        # Implied value = current_price * (median_peer / target_val)
        if target_val != 0:
            implied = current_price * (median_peer / target_val)
            implied_values.append(implied)
            multiple_details[mult_name] = {
                "target": round(target_val, 2),
                "peer_median": round(median_peer, 2),
                "implied_value": round(implied, 2),
            }

    if not implied_values:
        return ValuationResult("Comps", 0, 0, {"error": "No valid multiples"})

    fair_value = np.mean(implied_values)
    confidence = min(0.5 + 0.1 * len(implied_values), 0.9)

    details = {
        "peers": [p.get("ticker", "?") for p in peer_data],
        "multiples": multiple_details,
        "current_price": current_price,
    }

    return ValuationResult("Comps", fair_value, confidence, details)


def _extract_multiples(info: dict) -> dict:
    return {
        "pe": info.get("trailingPE") or info.get("forwardPE"),
        "ev_ebitda": info.get("enterpriseToEbitda"),
        "ps": info.get("priceToSalesTrailing12Months"),
        "peg": info.get("pegRatio"),
        "pb": info.get("priceToBook"),
    }


def _multiple_key(name: str) -> str:
    mapping = {"P/E": "pe", "EV/EBITDA": "ev_ebitda", "P/S": "ps",
               "PEG": "peg", "P/B": "pb"}
    return mapping.get(name, name.lower())


# ── 3c. Analyst Price Targets ────────────────────────────────────────────────

def analyst_valuation(data: dict) -> ValuationResult:
    """Use analyst consensus price targets."""
    info = data.get("info", {})

    target_mean = info.get("targetMeanPrice")
    target_median = info.get("targetMedianPrice")
    target_high = info.get("targetHighPrice")
    target_low = info.get("targetLowPrice")
    num_analysts = info.get("numberOfAnalystOpinions", 0)

    if not target_mean and not target_median:
        return ValuationResult("Analyst Targets", 0, 0,
                               {"error": "No analyst target data"})

    fair_value = target_median or target_mean

    # Confidence scales with number of analysts
    if num_analysts >= 20:
        confidence = 0.85
    elif num_analysts >= 10:
        confidence = 0.70
    elif num_analysts >= 5:
        confidence = 0.55
    else:
        confidence = 0.35

    # Recommendation trend
    recommendations = data.get("recommendations", [])
    rec_summary = {}
    for rec in recommendations[-10:]:
        grade = str(rec.get("toGrade", rec.get("To Grade", ""))).lower()
        if not grade:
            continue
        if any(w in grade for w in ("buy", "outperform", "overweight")):
            rec_summary["buy"] = rec_summary.get("buy", 0) + 1
        elif any(w in grade for w in ("sell", "underperform", "underweight")):
            rec_summary["sell"] = rec_summary.get("sell", 0) + 1
        else:
            rec_summary["hold"] = rec_summary.get("hold", 0) + 1

    details = {
        "target_mean": target_mean,
        "target_median": target_median,
        "target_high": target_high,
        "target_low": target_low,
        "num_analysts": num_analysts,
        "recommendation_summary": rec_summary,
    }

    return ValuationResult("Analyst Targets", fair_value, confidence, details)


# ── 3d. Technical Analysis ───────────────────────────────────────────────────

def technical_valuation(data: dict) -> ValuationResult:
    """Simple technical analysis based on moving averages and RSI."""
    hist_raw = data.get("history", {})
    if not hist_raw or "Close" not in hist_raw:
        return ValuationResult("Technical", 0, 0, {"error": "No price history"})

    close_prices = [v for v in hist_raw["Close"] if v is not None]
    if len(close_prices) < 200:
        return ValuationResult("Technical", 0, 0, {"error": "Insufficient history"})

    prices = pd.Series(close_prices)
    current_price = prices.iloc[-1]

    # Moving averages
    sma_50 = prices.tail(50).mean()
    sma_200 = prices.tail(200).mean()
    ema_20 = prices.ewm(span=20, adjust=False).mean().iloc[-1]

    # RSI (14-day)
    delta = prices.diff().tail(15)
    gain = delta.where(delta > 0, 0.0).mean()
    loss = (-delta.where(delta < 0, 0.0)).mean()
    rs = gain / loss if loss != 0 else 100
    rsi = 100 - (100 / (1 + rs))

    # MACD
    ema_12 = prices.ewm(span=12, adjust=False).mean().iloc[-1]
    ema_26 = prices.ewm(span=26, adjust=False).mean().iloc[-1]
    macd = ema_12 - ema_26

    # Bollinger Bands (20-day)
    bb_sma = prices.tail(20).mean()
    bb_std = prices.tail(20).std()
    bb_upper = bb_sma + 2 * bb_std
    bb_lower = bb_sma - 2 * bb_std

    # Score signals
    signals = []
    if current_price > sma_50 > sma_200:
        signals.append(("Golden Cross / Uptrend", 1))
    elif current_price < sma_50 < sma_200:
        signals.append(("Death Cross / Downtrend", -1))
    else:
        signals.append(("Mixed trend", 0))

    if rsi < 30:
        signals.append(("RSI oversold", 1))
    elif rsi > 70:
        signals.append(("RSI overbought", -1))
    else:
        signals.append(("RSI neutral", 0))

    if macd > 0:
        signals.append(("MACD bullish", 1))
    else:
        signals.append(("MACD bearish", -1))

    if current_price < bb_lower:
        signals.append(("Below Bollinger lower band", 1))
    elif current_price > bb_upper:
        signals.append(("Above Bollinger upper band", -1))
    else:
        signals.append(("Within Bollinger bands", 0))

    # Compute technical fair value as an adjusted price
    score = sum(s[1] for s in signals) / len(signals)  # -1 to 1
    adjustment = 1 + score * 0.10  # +/- 10% max adjustment
    fair_value = current_price * adjustment

    # Confidence is moderate for pure technicals
    confidence = 0.50

    details = {
        "current_price": round(current_price, 2),
        "sma_50": round(sma_50, 2),
        "sma_200": round(sma_200, 2),
        "ema_20": round(ema_20, 2),
        "rsi": round(rsi, 2),
        "macd": round(macd, 2),
        "bollinger_upper": round(bb_upper, 2),
        "bollinger_lower": round(bb_lower, 2),
        "signals": signals,
        "score": round(score, 3),
    }

    return ValuationResult("Technical", fair_value, confidence, details)


# ── 3e. Sentiment Analysis ───────────────────────────────────────────────────

def sentiment_valuation(data: dict, ticker: str) -> ValuationResult:
    """Sentiment analysis from news headlines and key metrics."""
    info = data.get("info", {})
    current_price = info.get("currentPrice") or info.get("regularMarketPrice", 0)
    if not current_price:
        return ValuationResult("Sentiment", 0, 0, {"error": "No current price"})

    scores = []

    # 1) News headline sentiment (simple keyword-based)
    headlines = fetch_news_sentiment(ticker)
    positive_words = {"beat", "surge", "jump", "gain", "rise", "profit", "growth",
                      "upgrade", "outperform", "rally", "record", "strong", "bullish",
                      "buy", "exceed", "boom", "soar", "positive", "optimistic"}
    negative_words = {"miss", "fall", "drop", "loss", "decline", "downgrade",
                      "underperform", "crash", "weak", "bearish", "sell", "cut",
                      "warn", "negative", "pessimistic", "layoff", "recall", "lawsuit"}

    if headlines:
        pos_count = 0
        neg_count = 0
        for h in headlines:
            text = ((h.get("title") or "") + " " + (h.get("description") or "")).lower()
            pos_count += sum(1 for w in positive_words if w in text)
            neg_count += sum(1 for w in negative_words if w in text)
        total = pos_count + neg_count
        if total > 0:
            news_score = (pos_count - neg_count) / total  # -1 to 1
            scores.append(news_score)

    # 2) Fundamental sentiment signals
    profit_margin = info.get("profitMargins")
    if profit_margin is not None:
        if profit_margin > 0.15:
            scores.append(0.5)
        elif profit_margin > 0:
            scores.append(0.2)
        else:
            scores.append(-0.5)

    revenue_growth = info.get("revenueGrowth")
    if revenue_growth is not None:
        if revenue_growth > 0.15:
            scores.append(0.5)
        elif revenue_growth > 0:
            scores.append(0.2)
        else:
            scores.append(-0.3)

    earnings_growth = info.get("earningsGrowth")
    if earnings_growth is not None:
        if earnings_growth > 0.15:
            scores.append(0.5)
        elif earnings_growth > 0:
            scores.append(0.2)
        else:
            scores.append(-0.3)

    # 3) Insider/institutional sentiment
    insider_pct = info.get("heldPercentInsiders", 0) or 0
    inst_pct = info.get("heldPercentInstitutions", 0) or 0
    if inst_pct > 0.7:
        scores.append(0.3)
    if insider_pct > 0.10:
        scores.append(0.2)

    if not scores:
        return ValuationResult("Sentiment", 0, 0, {"error": "No sentiment data"})

    avg_score = np.mean(scores)  # -1 to 1
    adjustment = 1 + avg_score * 0.08  # +/- 8% max
    fair_value = current_price * adjustment

    confidence = min(0.3 + 0.1 * len(scores), 0.65)

    details = {
        "current_price": current_price,
        "headline_count": len(headlines),
        "sentiment_score": round(avg_score, 3),
        "individual_scores": [round(s, 3) for s in scores],
        "adjustment_factor": round(adjustment, 4),
    }

    return ValuationResult("Sentiment", fair_value, confidence, details)


# ── 3f. Seasonal Analysis ────────────────────────────────────────────────────

def seasonal_valuation(data: dict) -> ValuationResult:
    """Seasonal analysis based on historical monthly return patterns."""
    hist_raw = data.get("history", {})
    if not hist_raw or "Close" not in hist_raw:
        return ValuationResult("Seasonal", 0, 0, {"error": "No price history"})

    dates = hist_raw.get("Date", [])
    closes = hist_raw.get("Close", [])

    if len(closes) < 252:
        return ValuationResult("Seasonal", 0, 0,
                               {"error": "Need at least 1 year of history"})

    # Build a DataFrame of monthly returns
    df = pd.DataFrame({"date": pd.to_datetime(dates), "close": closes})
    df = df.dropna(subset=["close"])
    df = df.set_index("date").sort_index()
    monthly = df["close"].resample("ME").last().dropna()
    monthly_ret = monthly.pct_change().dropna()

    if len(monthly_ret) < 12:
        return ValuationResult("Seasonal", 0, 0,
                               {"error": "Not enough monthly data"})

    monthly_ret_df = pd.DataFrame({
        "month": monthly_ret.index.month,
        "return": monthly_ret.values,
    })

    # Average return and win-rate for each calendar month
    month_stats = monthly_ret_df.groupby("month")["return"].agg(
        avg_return="mean", win_rate=lambda x: (x > 0).mean(), count="count"
    )

    current_month = datetime.datetime.now().month
    # Look-ahead window: average return over the next 3 months historically
    forward_months = [(current_month + i - 1) % 12 + 1 for i in range(1, 4)]
    fwd_returns = [
        month_stats.loc[m, "avg_return"]
        for m in forward_months
        if m in month_stats.index
    ]

    if not fwd_returns:
        return ValuationResult("Seasonal", 0, 0,
                               {"error": "Seasonal data incomplete"})

    expected_3m_return = sum(fwd_returns)

    info = data.get("info", {})
    current_price = info.get("currentPrice") or info.get("regularMarketPrice", 0)
    if not current_price:
        return ValuationResult("Seasonal", 0, 0, {"error": "No current price"})

    fair_value = current_price * (1 + expected_3m_return)

    # Confidence based on number of years of data
    years_of_data = len(monthly_ret) / 12
    confidence = min(0.3 + 0.1 * years_of_data, 0.65)

    # Build month-name stats for details
    month_names = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
                   "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
    monthly_detail = {}
    for m in range(1, 13):
        if m in month_stats.index:
            monthly_detail[month_names[m - 1]] = {
                "avg_return": round(month_stats.loc[m, "avg_return"], 4),
                "win_rate": round(month_stats.loc[m, "win_rate"], 4),
                "samples": int(month_stats.loc[m, "count"]),
            }

    details = {
        "current_price": current_price,
        "expected_3m_return": round(expected_3m_return, 4),
        "forward_months": [month_names[m - 1] for m in forward_months],
        "monthly_stats": monthly_detail,
        "years_of_data": round(years_of_data, 1),
    }

    return ValuationResult("Seasonal", fair_value, confidence, details)


# ── 3g. Historical P/E Valuation ────────────────────────────────────────────

def historical_pe_valuation(data: dict) -> ValuationResult:
    """Value stock based on its historical P/E range and current earnings."""
    info = data.get("info", {})
    current_price = info.get("currentPrice") or info.get("regularMarketPrice", 0)
    trailing_eps = info.get("trailingEps")
    forward_eps = info.get("forwardEps")

    if not current_price or not trailing_eps or trailing_eps <= 0:
        return ValuationResult("Historical P/E", 0, 0,
                               {"error": "No earnings data for P/E valuation"})

    current_pe = current_price / trailing_eps

    # Gather historical P/E data points from price history and earnings
    hist_raw = data.get("history", {})
    closes = hist_raw.get("Close", [])

    # Estimate historical P/E range from 5Y average PE
    five_yr_avg_pe = info.get("fiveYearAvgDividendYield")  # not right, use trailingPE
    sector_pe = info.get("sectorPE")

    # Use a range: industry/sector averages + own historical
    pe_estimates = [current_pe]

    # Forward PE if available
    if forward_eps and forward_eps > 0:
        forward_pe = current_price / forward_eps
        pe_estimates.append(forward_pe)

    # Sector average (approximate from peers if not directly available)
    trailing_pe = info.get("trailingPE")
    forward_pe_info = info.get("forwardPE")
    if trailing_pe and trailing_pe > 0:
        pe_estimates.append(trailing_pe)
    if forward_pe_info and forward_pe_info > 0:
        pe_estimates.append(forward_pe_info)

    # Historical average P/E approximation: use 5Y price range with current EPS
    if len(closes) > 252:
        five_yr_prices = closes[-1260:] if len(closes) >= 1260 else closes
        avg_price = np.mean(five_yr_prices)
        hist_avg_pe = avg_price / trailing_eps if trailing_eps > 0 else None
        if hist_avg_pe and 5 < hist_avg_pe < 100:
            pe_estimates.append(hist_avg_pe)

    avg_pe = np.median(pe_estimates)
    # Clip to reasonable range
    avg_pe = np.clip(avg_pe, 8, 60)

    # Fair value = historical average PE * current/forward EPS
    eps_to_use = forward_eps if forward_eps and forward_eps > 0 else trailing_eps
    fair_value = avg_pe * eps_to_use

    # Confidence based on data richness
    confidence = 0.50
    if len(pe_estimates) >= 3:
        confidence += 0.10
    if forward_eps and forward_eps > 0:
        confidence += 0.10
    if len(closes) > 1000:
        confidence += 0.10
    confidence = min(confidence, 0.80)

    details = {
        "current_pe": round(current_pe, 2),
        "historical_avg_pe": round(avg_pe, 2),
        "trailing_eps": round(trailing_eps, 2),
        "forward_eps": round(forward_eps, 2) if forward_eps else None,
        "eps_used": round(eps_to_use, 2),
        "pe_data_points": len(pe_estimates),
        "current_price": round(current_price, 2),
    }

    return ValuationResult("Historical P/E", fair_value, confidence, details)


# ── 3h. Quality Score ───────────────────────────────────────────────────────

def quality_score_analysis(data: dict) -> dict:
    """Calculate a composite quality score (0-100) across multiple dimensions."""
    info = data.get("info", {})
    scores = {}

    # 1) Profitability (25%)
    gross_margin = info.get("grossMargins", 0) or 0
    op_margin = info.get("operatingMargins", 0) or 0
    net_margin = info.get("profitMargins", 0) or 0
    roe = info.get("returnOnEquity", 0) or 0

    prof_score = 0
    if gross_margin >= cfg.QUALITY_THRESHOLDS["gross_margin_target"]:
        prof_score += 30
    elif gross_margin > 0.20:
        prof_score += 15
    if op_margin > 0.15:
        prof_score += 25
    elif op_margin > 0.05:
        prof_score += 12
    if net_margin > 0.10:
        prof_score += 20
    elif net_margin > 0:
        prof_score += 10
    if roe >= cfg.QUALITY_THRESHOLDS["roe_target"]:
        prof_score += 25
    elif roe > 0.10:
        prof_score += 12
    scores["profitability"] = min(prof_score, 100)

    # 2) Growth (20%)
    rev_growth = info.get("revenueGrowth", 0) or 0
    earn_growth = info.get("earningsGrowth", 0) or 0
    growth_score = 0
    if rev_growth >= cfg.QUALITY_THRESHOLDS["revenue_growth_target"]:
        growth_score += 50
    elif rev_growth > 0.05:
        growth_score += 25
    elif rev_growth > 0:
        growth_score += 10
    if earn_growth > 0.15:
        growth_score += 50
    elif earn_growth > 0.05:
        growth_score += 25
    elif earn_growth > 0:
        growth_score += 10
    scores["growth"] = min(growth_score, 100)

    # 3) Competitive Moat (20%)
    moat_score = 0
    if gross_margin > 0.50:
        moat_score += 40  # High margin = pricing power
    elif gross_margin > 0.35:
        moat_score += 20
    market_cap = info.get("marketCap", 0) or 0
    if market_cap > 100e9:
        moat_score += 30  # Large cap = established
    elif market_cap > 10e9:
        moat_score += 15
    # Consistent revenue (proxy: low beta)
    beta = info.get("beta", 1.0) or 1.0
    if beta < 0.8:
        moat_score += 30
    elif beta < 1.2:
        moat_score += 15
    scores["moat"] = min(moat_score, 100)

    # 4) Financial Health (15%)
    health_score = 0
    current_ratio = info.get("currentRatio", 0) or 0
    debt_equity = info.get("debtToEquity", 0) or 0
    if debt_equity > 10:
        debt_equity = debt_equity / 100.0  # Normalize if in %

    if current_ratio >= 1.5:
        health_score += 35
    elif current_ratio >= 1.0:
        health_score += 18
    if debt_equity <= cfg.QUALITY_THRESHOLDS["debt_equity_target"]:
        health_score += 35
    elif debt_equity <= 1.0:
        health_score += 18
    total_cash = info.get("totalCash", 0) or 0
    total_debt = info.get("totalDebt", 0) or 0
    if total_cash > total_debt:
        health_score += 30
    elif total_debt > 0 and total_cash / total_debt > 0.5:
        health_score += 15
    scores["financial_health"] = min(health_score, 100)

    # 5) Management Effectiveness (10%)
    mgmt_score = 0
    roa = info.get("returnOnAssets", 0) or 0
    if roe > 0.20:
        mgmt_score += 40
    elif roe > 0.12:
        mgmt_score += 20
    if roa > 0.10:
        mgmt_score += 30
    elif roa > 0.05:
        mgmt_score += 15
    insider_pct = info.get("heldPercentInsiders", 0) or 0
    if 0.01 < insider_pct < 0.30:
        mgmt_score += 30  # Skin in the game but not too concentrated
    scores["management"] = min(mgmt_score, 100)

    # 6) Innovation (10%)
    innov_score = 0
    # Use R&D from financials if available
    financials = data.get("financials", {})
    is_data = financials.get("income_stmt", {})
    if is_data:
        is_df = pd.DataFrame(is_data)
        for label in ("Research Development", "ResearchAndDevelopment",
                       "Research And Development Expenses"):
            if label in is_df.index:
                rd_vals = [v for v in is_df.loc[label].values
                           if v is not None and not np.isnan(float(v))]
                if rd_vals:
                    rd = abs(float(rd_vals[0]))
                    revenue = info.get("totalRevenue", 0) or 0
                    if revenue > 0:
                        rd_ratio = rd / revenue
                        if rd_ratio >= cfg.QUALITY_THRESHOLDS["rd_revenue_target"]:
                            innov_score += 60
                        elif rd_ratio > 0.05:
                            innov_score += 30
                break
    # Tech sector bonus
    sector = (info.get("sector", "") or "").lower()
    if sector in ("technology", "healthcare", "communication services"):
        innov_score += 40
    scores["innovation"] = min(innov_score, 100)

    # Weighted composite
    composite = sum(
        scores.get(dim, 0) * w
        for dim, w in cfg.QUALITY_WEIGHTS.items()
    )

    # Grade
    if composite >= 80:
        grade = "A"
    elif composite >= 65:
        grade = "B"
    elif composite >= 50:
        grade = "C"
    elif composite >= 35:
        grade = "D"
    else:
        grade = "F"

    return {
        "scores": scores,
        "composite": round(composite, 1),
        "grade": grade,
    }


# ── 3i. Scenario Analysis ──────────────────────────────────────────────────

def scenario_analysis(data: dict, results: list) -> dict:
    """Generate bear / base / bull fair value scenarios."""
    info = data.get("info", {})
    current_price = info.get("currentPrice") or info.get("regularMarketPrice", 0)

    # Collect valid fair values from all models
    valid_fvs = [r.fair_value for r in results if r.fair_value > 0]
    if not valid_fvs or not current_price:
        return {"bear": 0, "base": 0, "bull": 0, "weighted": 0,
                "details": {"error": "Insufficient data"}}

    base_fv = np.median(valid_fvs)

    # Bear: 25th percentile of models, then apply 15% haircut
    bear_fv = np.percentile(valid_fvs, 25) * 0.85

    # Bull: 75th percentile of models, then apply 15% premium
    bull_fv = np.percentile(valid_fvs, 75) * 1.15

    # Weighted scenario value
    w = cfg.SCENARIO_WEIGHTS
    weighted = (bear_fv * w["bear"] + base_fv * w["base"] + bull_fv * w["bull"])

    return {
        "bear": round(bear_fv, 2),
        "base": round(base_fv, 2),
        "bull": round(bull_fv, 2),
        "weighted": round(weighted, 2),
        "bear_upside": round((bear_fv - current_price) / current_price, 4)
            if current_price else 0,
        "base_upside": round((base_fv - current_price) / current_price, 4)
            if current_price else 0,
        "bull_upside": round((bull_fv - current_price) / current_price, 4)
            if current_price else 0,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 4 — Composite Valuation
# ═══════════════════════════════════════════════════════════════════════════════

def composite_valuation(results: list, weights: dict) -> dict:
    """Combine multiple valuation results into a single weighted estimate."""
    method_map = {
        "DCF": "dcf",
        "Comps": "comps",
        "Analyst Targets": "analyst",
        "Technical": "technical",
        "Sentiment": "sentiment",
        "Seasonal": "seasonal",
        "Historical P/E": "historical",
    }

    weighted_sum = 0.0
    total_weight = 0.0
    breakdown = {}

    for r in results:
        key = method_map.get(r.method, r.method.lower())
        w = weights.get(key, 0)

        if r.fair_value > 0 and r.confidence > 0:
            effective_weight = w * r.confidence
            weighted_sum += r.fair_value * effective_weight
            total_weight += effective_weight
            breakdown[r.method] = {
                "fair_value": round(r.fair_value, 2),
                "confidence": round(r.confidence, 2),
                "weight": w,
                "effective_weight": round(effective_weight, 4),
            }
        else:
            breakdown[r.method] = {
                "fair_value": 0,
                "confidence": 0,
                "weight": w,
                "effective_weight": 0,
                "note": r.details.get("error", "No data"),
            }

    composite_value = weighted_sum / total_weight if total_weight > 0 else 0

    return {
        "composite_fair_value": round(composite_value, 2),
        "total_effective_weight": round(total_weight, 4),
        "breakdown": breakdown,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 5 — Macro Environment
# ═══════════════════════════════════════════════════════════════════════════════

def fetch_macro_environment() -> dict:
    """Fetch key macroeconomic indicators from FRED."""
    print("  [FRED] Fetching macro indicators...")
    macro = {}
    for name, series_id in cfg.FRED_SERIES.items():
        val = fetch_fred_series(series_id)
        macro[name] = val
    return macro


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 6 — Excel Report Generation
# ═══════════════════════════════════════════════════════════════════════════════

def _style_header(ws, row, max_col):
    """Apply header styling to a row."""
    hdr_fill = PatternFill(start_color=cfg.COLOR_SCHEME["header_bg"],
                           end_color=cfg.COLOR_SCHEME["header_bg"],
                           fill_type="solid")
    hdr_font = Font(color=cfg.COLOR_SCHEME["header_font"], bold=True, size=11)
    thin_border = Border(
        bottom=Side(style="thin", color=cfg.COLOR_SCHEME["light_border"])
    )
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border


def _fmt_number(value, fmt_type="number"):
    """Format a number for display."""
    if value is None:
        return "N/A"
    try:
        value = float(value)
    except (TypeError, ValueError):
        return str(value)
    if fmt_type == "currency":
        return f"${value:,.2f}"
    if fmt_type == "percent":
        return f"{value:.2%}"
    if fmt_type == "large_currency":
        if abs(value) >= 1e12:
            return f"${value / 1e12:,.2f}T"
        if abs(value) >= 1e9:
            return f"${value / 1e9:,.2f}B"
        if abs(value) >= 1e6:
            return f"${value / 1e6:,.2f}M"
        return f"${value:,.0f}"
    if fmt_type == "ratio":
        return f"{value:.2f}x"
    return f"{value:,.2f}"


def generate_excel_report(ticker: str, data: dict, results: list,
                          composite: dict, macro: dict,
                          quality_data: dict = None,
                          scenario_data: dict = None) -> str:
    """Generate a comprehensive Excel report with ~20 sheets."""
    os.makedirs(cfg.OUTPUT_DIR, exist_ok=True)
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{ticker}_valuation_{timestamp}.xlsx"
    filepath = os.path.join(cfg.OUTPUT_DIR, filename)

    wb = Workbook()
    info = data.get("info", {})
    current_price = info.get("currentPrice") or info.get("regularMarketPrice", 0)

    # ── Sheet 1: Executive Summary ──
    ws = wb.active
    ws.title = "Executive Summary"
    ws.sheet_properties.tabColor = cfg.COLOR_SCHEME["header_bg"]

    # Title
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = f"Stock Valuation Report — {ticker}"
    title_cell.font = Font(size=18, bold=True, color=cfg.COLOR_SCHEME["header_bg"])
    title_cell.alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:G2")
    ws["A2"].value = (f"{info.get('shortName', ticker)}  |  "
                      f"{info.get('sector', 'N/A')}  |  "
                      f"{info.get('industry', 'N/A')}")
    ws["A2"].font = Font(size=11, italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:G3")
    ws["A3"].value = f"Report generated: {datetime.datetime.now():%Y-%m-%d %H:%M}"
    ws["A3"].font = Font(size=9, color=cfg.COLOR_SCHEME["neutral"])
    ws["A3"].alignment = Alignment(horizontal="center")

    # Key metrics box
    row = 5
    headers = ["Metric", "Value"]
    for c, h in enumerate(headers, 1):
        ws.cell(row=row, column=c, value=h)
    _style_header(ws, row, len(headers))

    key_metrics = [
        ("Current Price", _fmt_number(current_price, "currency")),
        ("Composite Fair Value", _fmt_number(composite["composite_fair_value"], "currency")),
        ("Upside / Downside",
         _fmt_number((composite["composite_fair_value"] - current_price) / current_price
                     if current_price else 0, "percent")),
        ("Market Cap", _fmt_number(info.get("marketCap"), "large_currency")),
        ("52-Week High", _fmt_number(info.get("fiftyTwoWeekHigh"), "currency")),
        ("52-Week Low", _fmt_number(info.get("fiftyTwoWeekLow"), "currency")),
        ("Beta", _fmt_number(info.get("beta"), "ratio")),
        ("Dividend Yield", _fmt_number(info.get("dividendYield"), "percent")),
        ("P/E (Trailing)", _fmt_number(info.get("trailingPE"), "ratio")),
        ("EPS (TTM)", _fmt_number(info.get("trailingEps"), "currency")),
    ]

    pos_font = Font(color=cfg.COLOR_SCHEME["positive"], bold=True)
    neg_font = Font(color=cfg.COLOR_SCHEME["negative"], bold=True)
    alt_fill = PatternFill(start_color=cfg.COLOR_SCHEME["alt_row"],
                           end_color=cfg.COLOR_SCHEME["alt_row"], fill_type="solid")

    for i, (metric, value) in enumerate(key_metrics):
        r = row + 1 + i
        ws.cell(row=r, column=1, value=metric)
        cell = ws.cell(row=r, column=2, value=value)
        if i % 2 == 0:
            ws.cell(row=r, column=1).fill = alt_fill
            cell.fill = alt_fill
        # Highlight upside/downside
        if metric == "Upside / Downside" and current_price > 0:
            ratio = (composite["composite_fair_value"] - current_price) / current_price
            cell.font = pos_font if ratio >= 0 else neg_font

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18

    # Signal / verdict
    r = row + len(key_metrics) + 2
    ws.merge_cells(f"A{r}:B{r}")
    if current_price > 0:
        pct_diff = (composite["composite_fair_value"] - current_price) / current_price
        if pct_diff > 0.15:
            verdict = "UNDERVALUED — Consider Buying"
            v_color = cfg.COLOR_SCHEME["positive"]
        elif pct_diff < -0.15:
            verdict = "OVERVALUED — Consider Selling"
            v_color = cfg.COLOR_SCHEME["negative"]
        else:
            verdict = "FAIRLY VALUED — Hold"
            v_color = cfg.COLOR_SCHEME["neutral"]
    else:
        verdict = "INSUFFICIENT DATA"
        v_color = cfg.COLOR_SCHEME["neutral"]

    vc = ws.cell(row=r, column=1, value=verdict)
    vc.font = Font(size=14, bold=True, color=v_color)
    vc.alignment = Alignment(horizontal="center")

    # ── Sheet 2: Valuation Breakdown ──
    ws2 = wb.create_sheet("Valuation Breakdown")
    ws2.sheet_properties.tabColor = "4472C4"

    ws2.merge_cells("A1:F1")
    ws2["A1"].value = "Valuation Method Comparison"
    ws2["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

    row = 3
    headers = ["Method", "Fair Value", "Confidence", "Weight",
               "Effective Weight", "Notes"]
    for c, h in enumerate(headers, 1):
        ws2.cell(row=row, column=c, value=h)
    _style_header(ws2, row, len(headers))

    for i, (method, bd) in enumerate(composite["breakdown"].items()):
        r = row + 1 + i
        ws2.cell(row=r, column=1, value=method)
        ws2.cell(row=r, column=2, value=_fmt_number(bd["fair_value"], "currency"))
        ws2.cell(row=r, column=3, value=_fmt_number(bd["confidence"], "percent"))
        ws2.cell(row=r, column=4, value=_fmt_number(bd["weight"], "percent"))
        ws2.cell(row=r, column=5, value=_fmt_number(bd["effective_weight"], "percent"))
        ws2.cell(row=r, column=6, value=bd.get("note", ""))
        if i % 2 == 0:
            for c in range(1, 7):
                ws2.cell(row=r, column=c).fill = alt_fill

    # Summary row
    r = row + 1 + len(composite["breakdown"])
    ws2.cell(row=r + 1, column=1, value="COMPOSITE FAIR VALUE").font = Font(bold=True)
    ws2.cell(row=r + 1, column=2,
             value=_fmt_number(composite["composite_fair_value"], "currency"))
    ws2.cell(row=r + 1, column=2).font = Font(bold=True, size=12)

    for c in range(1, 7):
        ws2.column_dimensions[get_column_letter(c)].width = 18

    # ── Sheet 3: DCF Detail ──
    ws3 = wb.create_sheet("DCF Analysis")
    ws3.sheet_properties.tabColor = "00B050"

    dcf_result = next((r for r in results if r.method == "DCF"), None)
    if dcf_result and dcf_result.fair_value > 0:
        d = dcf_result.details
        ws3["A1"].value = "Discounted Cash Flow Analysis"
        ws3["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

        row = 3
        assumptions = [
            ("Latest Free Cash Flow", _fmt_number(d.get("latest_fcf"), "large_currency")),
            ("FCF Growth Rate", _fmt_number(d.get("fcf_growth_rate"), "percent")),
            ("WACC (Discount Rate)", _fmt_number(d.get("wacc"), "percent")),
            ("Cost of Equity", _fmt_number(d.get("cost_of_equity"), "percent")),
            ("Cost of Debt", _fmt_number(d.get("cost_of_debt"), "percent")),
            ("Beta", _fmt_number(d.get("beta"))),
            ("Risk-Free Rate", _fmt_number(d.get("risk_free_rate"), "percent")),
            ("Terminal Value", _fmt_number(d.get("terminal_value"), "large_currency")),
            ("Enterprise Value", _fmt_number(d.get("enterprise_value"), "large_currency")),
            ("Equity Value", _fmt_number(d.get("equity_value"), "large_currency")),
            ("Shares Outstanding", _fmt_number(d.get("shares_outstanding"), "large_currency")),
            ("DCF Fair Value / Share", _fmt_number(dcf_result.fair_value, "currency")),
            ("Margin of Safety", _fmt_number(d.get("margin_of_safety"), "percent")),
            ("Buy Price (w/ MoS)", _fmt_number(d.get("buy_price"), "currency")),
        ]

        headers = ["Assumption", "Value"]
        for c, h in enumerate(headers, 1):
            ws3.cell(row=row, column=c, value=h)
        _style_header(ws3, row, 2)

        for i, (label, val) in enumerate(assumptions):
            r = row + 1 + i
            ws3.cell(row=r, column=1, value=label)
            ws3.cell(row=r, column=2, value=val)
            if i % 2 == 0:
                ws3.cell(row=r, column=1).fill = alt_fill
                ws3.cell(row=r, column=2).fill = alt_fill

        # Projected FCFs
        proj_row = row + len(assumptions) + 2
        ws3.cell(row=proj_row, column=1, value="Projected Free Cash Flows").font = Font(
            bold=True)
        proj_fcfs = d.get("projected_fcfs", [])
        for i, fcf in enumerate(proj_fcfs):
            ws3.cell(row=proj_row + 1 + i, column=1, value=f"Year {i + 1}")
            ws3.cell(row=proj_row + 1 + i, column=2,
                     value=_fmt_number(fcf, "large_currency"))

        # DCF Sensitivity Matrix (WACC vs Terminal Growth Rate)
        sens_row = proj_row + len(proj_fcfs) + 3
        ws3.cell(row=sens_row, column=1,
                 value="Sensitivity Analysis — Fair Value per Share").font = Font(bold=True)
        sens_row += 1

        wacc_base = d.get("wacc", 0.10)
        tgr_base = cfg.DCF_DEFAULTS["terminal_growth_rate"]
        wacc_steps = cfg.DCF_SENSITIVITY["wacc_steps"]
        tgr_steps = cfg.DCF_SENSITIVITY["tgr_steps"]

        # Column headers (TGR values)
        ws3.cell(row=sens_row, column=1, value="WACC \\ TGR")
        for j, tgr_off in enumerate(tgr_steps):
            tgr_val = tgr_base + tgr_off
            ws3.cell(row=sens_row, column=2 + j,
                     value=f"{tgr_val:.1%}")
        _style_header(ws3, sens_row, 1 + len(tgr_steps))

        shares = d.get("shares_outstanding", 1)
        cash = data.get("info", {}).get("totalCash", 0) or 0
        total_debt = data.get("info", {}).get("totalDebt", 0) or 0

        good_fill = PatternFill(start_color=cfg.COLOR_SCHEME["good_bg"],
                                end_color=cfg.COLOR_SCHEME["good_bg"],
                                fill_type="solid")
        bad_fill = PatternFill(start_color=cfg.COLOR_SCHEME["bad_bg"],
                               end_color=cfg.COLOR_SCHEME["bad_bg"],
                               fill_type="solid")
        warn_fill_bg = PatternFill(start_color=cfg.COLOR_SCHEME["warn_bg"],
                                   end_color=cfg.COLOR_SCHEME["warn_bg"],
                                   fill_type="solid")

        for i, wacc_off in enumerate(wacc_steps):
            r = sens_row + 1 + i
            w = wacc_base + wacc_off
            ws3.cell(row=r, column=1, value=f"{w:.2%}").font = Font(bold=True)
            for j, tgr_off in enumerate(tgr_steps):
                tg = tgr_base + tgr_off
                if w > tg:
                    # Recompute terminal value & fair value
                    t_fcf = proj_fcfs[-1] * (1 + tg) if proj_fcfs else 0
                    tv = t_fcf / (w - tg)
                    pv_t = tv / ((1 + w) ** len(proj_fcfs))
                    pv_f = sum(f / ((1 + w) ** yr)
                               for yr, f in enumerate(proj_fcfs, 1))
                    ev = pv_f + pv_t
                    eq = ev + cash - total_debt
                    fv = eq / shares if shares > 0 else 0
                else:
                    fv = 0  # invalid when WACC <= TGR

                cell = ws3.cell(row=r, column=2 + j,
                                value=_fmt_number(fv, "currency"))
                # Color code relative to current price
                current_price_val = data.get("info", {}).get(
                    "currentPrice") or data.get("info", {}).get(
                    "regularMarketPrice", 0)
                if current_price_val and fv > 0:
                    pct = (fv - current_price_val) / current_price_val
                    if pct > 0.15:
                        cell.fill = good_fill
                    elif pct < -0.15:
                        cell.fill = bad_fill
                    else:
                        cell.fill = warn_fill_bg

        ws3.column_dimensions["A"].width = 28
        ws3.column_dimensions["B"].width = 14
        for j in range(len(tgr_steps)):
            ws3.column_dimensions[get_column_letter(3 + j)].width = 14
    else:
        ws3["A1"].value = "DCF Analysis — Insufficient Data"
        ws3["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["negative"])
        if dcf_result:
            ws3["A3"].value = dcf_result.details.get("error", "")

    # ── Sheet 4: Technical Analysis ──
    ws4 = wb.create_sheet("Technical Analysis")
    ws4.sheet_properties.tabColor = "FFC000"

    tech_result = next((r for r in results if r.method == "Technical"), None)
    if tech_result and tech_result.fair_value > 0:
        d = tech_result.details
        ws4["A1"].value = "Technical Analysis"
        ws4["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

        row = 3
        indicators = [
            ("Current Price", _fmt_number(d.get("current_price"), "currency")),
            ("50-Day SMA", _fmt_number(d.get("sma_50"), "currency")),
            ("200-Day SMA", _fmt_number(d.get("sma_200"), "currency")),
            ("20-Day EMA", _fmt_number(d.get("ema_20"), "currency")),
            ("RSI (14)", _fmt_number(d.get("rsi"))),
            ("MACD", _fmt_number(d.get("macd"), "currency")),
            ("Bollinger Upper", _fmt_number(d.get("bollinger_upper"), "currency")),
            ("Bollinger Lower", _fmt_number(d.get("bollinger_lower"), "currency")),
            ("Technical Score", _fmt_number(d.get("score"))),
            ("Technical Fair Value", _fmt_number(tech_result.fair_value, "currency")),
        ]

        headers = ["Indicator", "Value"]
        for c, h in enumerate(headers, 1):
            ws4.cell(row=row, column=c, value=h)
        _style_header(ws4, row, 2)

        for i, (label, val) in enumerate(indicators):
            r = row + 1 + i
            ws4.cell(row=r, column=1, value=label)
            ws4.cell(row=r, column=2, value=val)
            if i % 2 == 0:
                ws4.cell(row=r, column=1).fill = alt_fill
                ws4.cell(row=r, column=2).fill = alt_fill

        # Signals
        sig_row = row + len(indicators) + 2
        ws4.cell(row=sig_row, column=1, value="Signals").font = Font(bold=True)
        for i, (signal, direction) in enumerate(d.get("signals", [])):
            r = sig_row + 1 + i
            ws4.cell(row=r, column=1, value=signal)
            label = "Bullish" if direction > 0 else ("Bearish" if direction < 0 else "Neutral")
            cell = ws4.cell(row=r, column=2, value=label)
            if direction > 0:
                cell.font = pos_font
            elif direction < 0:
                cell.font = neg_font

        ws4.column_dimensions["A"].width = 24
        ws4.column_dimensions["B"].width = 16
    else:
        ws4["A1"].value = "Technical Analysis — Insufficient Data"
        ws4["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["negative"])

    # ── Sheet 5: Macro Environment ──
    ws5 = wb.create_sheet("Macro Environment")
    ws5.sheet_properties.tabColor = "7030A0"

    ws5["A1"].value = "Macroeconomic Environment"
    ws5["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

    row = 3
    headers = ["Indicator", "Latest Value"]
    for c, h in enumerate(headers, 1):
        ws5.cell(row=row, column=c, value=h)
    _style_header(ws5, row, 2)

    for i, (name, value) in enumerate(macro.items()):
        r = row + 1 + i
        ws5.cell(row=r, column=1, value=name)
        if value is not None:
            ws5.cell(row=r, column=2, value=round(value, 2))
        else:
            ws5.cell(row=r, column=2, value="N/A")
            ws5.cell(row=r, column=2).font = Font(color=cfg.COLOR_SCHEME["neutral"])
        if i % 2 == 0:
            ws5.cell(row=r, column=1).fill = alt_fill
            ws5.cell(row=r, column=2).fill = alt_fill

    ws5.column_dimensions["A"].width = 28
    ws5.column_dimensions["B"].width = 16

    # ── Sheet 6: Financial Ratios ──
    ws_ratios = wb.create_sheet("Financial Ratios")
    ws_ratios.sheet_properties.tabColor = "00B0F0"

    ws_ratios["A1"].value = "Financial Ratio Analysis"
    ws_ratios["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

    row = 3
    headers = ["Ratio", "Value", "Benchmark", "Status", "Explanation"]
    for c, h in enumerate(headers, 1):
        ws_ratios.cell(row=row, column=c, value=h)
    _style_header(ws_ratios, row, len(headers))

    good_fill = PatternFill(start_color=cfg.COLOR_SCHEME["good_bg"],
                            end_color=cfg.COLOR_SCHEME["good_bg"], fill_type="solid")
    warn_fill_r = PatternFill(start_color=cfg.COLOR_SCHEME["warn_bg"],
                              end_color=cfg.COLOR_SCHEME["warn_bg"], fill_type="solid")
    bad_fill_r = PatternFill(start_color=cfg.COLOR_SCHEME["bad_bg"],
                             end_color=cfg.COLOR_SCHEME["bad_bg"], fill_type="solid")

    # Map config ratio names to yfinance info keys
    ratio_keys = {
        "Current Ratio": "currentRatio",
        "Quick Ratio": "quickRatio",
        "Gross Margin": "grossMargins",
        "Operating Margin": "operatingMargins",
        "Net Margin": "profitMargins",
        "ROE": "returnOnEquity",
        "ROA": "returnOnAssets",
        "Debt / Equity": "debtToEquity",
        "PEG Ratio": "pegRatio",
    }

    ratio_row = row
    for ratio_name, bench in cfg.RATIO_BENCHMARKS.items():
        info_key = ratio_keys.get(ratio_name)
        if not info_key:
            continue
        raw_val = info.get(info_key)
        if raw_val is None:
            continue

        ratio_row += 1
        val = float(raw_val)
        # debtToEquity from yfinance is already a ratio (e.g. 1.5 = 150%)
        if info_key == "debtToEquity":
            val = val / 100.0 if val > 10 else val  # yfinance sometimes returns as %

        ws_ratios.cell(row=ratio_row, column=1, value=ratio_name)

        # Format value
        if ratio_name in ("Current Ratio", "Quick Ratio", "Cash Ratio",
                          "Debt / Equity", "Debt / Assets", "PEG Ratio",
                          "Asset Turnover", "Interest Coverage"):
            ws_ratios.cell(row=ratio_row, column=2, value=_fmt_number(val, "ratio"))
        else:
            ws_ratios.cell(row=ratio_row, column=2, value=_fmt_number(val, "percent"))

        # Determine status based on benchmark type
        if "good_below" in bench:
            # Lower is better
            bench_str = f"< {bench['good_below']}"
            if val <= bench["good_below"]:
                status = "Good"
                fill = good_fill
            elif val >= bench["warn_above"]:
                status = "Concern"
                fill = bad_fill_r
            else:
                status = "Watch"
                fill = warn_fill_r
        else:
            # Higher is better
            bench_str = f"> {bench['good']}"
            if val >= bench["good"]:
                status = "Good"
                fill = good_fill
            elif val <= bench["warn"]:
                status = "Concern"
                fill = bad_fill_r
            else:
                status = "Watch"
                fill = warn_fill_r

        ws_ratios.cell(row=ratio_row, column=3, value=bench_str)
        status_cell = ws_ratios.cell(row=ratio_row, column=4, value=status)
        explanation = cfg.RATIO_EXPLANATIONS.get(ratio_name, "")
        ws_ratios.cell(row=ratio_row, column=5, value=explanation)
        ws_ratios.cell(row=ratio_row, column=5).alignment = Alignment(wrap_text=True)
        for c in range(1, 6):
            ws_ratios.cell(row=ratio_row, column=c).fill = fill

    ws_ratios.column_dimensions["A"].width = 22
    ws_ratios.column_dimensions["B"].width = 14
    ws_ratios.column_dimensions["C"].width = 14
    ws_ratios.column_dimensions["D"].width = 12
    ws_ratios.column_dimensions["E"].width = 50

    # ── Sheet 7: Seasonal Analysis ──
    seasonal_result = next((r for r in results if r.method == "Seasonal"), None)
    ws_seas = wb.create_sheet("Seasonal Analysis")
    ws_seas.sheet_properties.tabColor = "92D050"

    if seasonal_result and seasonal_result.fair_value > 0:
        sd = seasonal_result.details
        ws_seas["A1"].value = "Seasonal Return Patterns"
        ws_seas["A1"].font = Font(size=14, bold=True,
                                  color=cfg.COLOR_SCHEME["header_bg"])

        row = 3
        headers = ["Month", "Avg Return", "Win Rate", "Samples"]
        for c, h in enumerate(headers, 1):
            ws_seas.cell(row=row, column=c, value=h)
        _style_header(ws_seas, row, len(headers))

        monthly_stats = sd.get("monthly_stats", {})
        for i, (month, stats) in enumerate(monthly_stats.items()):
            r = row + 1 + i
            ws_seas.cell(row=r, column=1, value=month)
            avg_ret = stats["avg_return"]
            ws_seas.cell(row=r, column=2, value=_fmt_number(avg_ret, "percent"))
            ws_seas.cell(row=r, column=3, value=_fmt_number(stats["win_rate"], "percent"))
            ws_seas.cell(row=r, column=4, value=stats["samples"])
            # Color code the return
            ret_cell = ws_seas.cell(row=r, column=2)
            if avg_ret > 0:
                ret_cell.font = pos_font
            elif avg_ret < 0:
                ret_cell.font = neg_font
            if i % 2 == 0:
                for c in range(1, 5):
                    ws_seas.cell(row=r, column=c).fill = alt_fill

        # Summary
        summary_row = row + len(monthly_stats) + 2
        ws_seas.cell(row=summary_row, column=1,
                     value="Forward 3-Month Outlook").font = Font(bold=True)
        ws_seas.cell(row=summary_row + 1, column=1, value="Forward Months")
        ws_seas.cell(row=summary_row + 1, column=2,
                     value=", ".join(sd.get("forward_months", [])))
        ws_seas.cell(row=summary_row + 2, column=1, value="Expected 3M Return")
        ws_seas.cell(row=summary_row + 2, column=2,
                     value=_fmt_number(sd.get("expected_3m_return"), "percent"))
        ws_seas.cell(row=summary_row + 3, column=1, value="Seasonal Fair Value")
        ws_seas.cell(row=summary_row + 3, column=2,
                     value=_fmt_number(seasonal_result.fair_value, "currency"))
        ws_seas.cell(row=summary_row + 4, column=1, value="Years of Data")
        ws_seas.cell(row=summary_row + 4, column=2,
                     value=sd.get("years_of_data"))

        ws_seas.column_dimensions["A"].width = 24
        ws_seas.column_dimensions["B"].width = 14
        ws_seas.column_dimensions["C"].width = 12
        ws_seas.column_dimensions["D"].width = 10
    else:
        ws_seas["A1"].value = "Seasonal Analysis — Insufficient Data"
        ws_seas["A1"].font = Font(size=14, bold=True,
                                  color=cfg.COLOR_SCHEME["negative"])
        if seasonal_result:
            ws_seas["A3"].value = seasonal_result.details.get("error", "")

    # ── Sheet 8: Company Profile ──
    ws6 = wb.create_sheet("Company Profile")
    ws6.sheet_properties.tabColor = "ED7D31"

    ws6["A1"].value = f"Company Profile — {info.get('shortName', ticker)}"
    ws6["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

    profile_fields = [
        ("Ticker", ticker),
        ("Company Name", info.get("shortName", "N/A")),
        ("Sector", info.get("sector", "N/A")),
        ("Industry", info.get("industry", "N/A")),
        ("Country", info.get("country", "N/A")),
        ("Exchange", info.get("exchange", "N/A")),
        ("Currency", info.get("currency", "N/A")),
        ("Website", info.get("website", "N/A")),
        ("Full-Time Employees", _fmt_number(info.get("fullTimeEmployees"))),
        ("Market Cap", _fmt_number(info.get("marketCap"), "large_currency")),
        ("Enterprise Value", _fmt_number(info.get("enterpriseValue"), "large_currency")),
        ("Revenue (TTM)", _fmt_number(info.get("totalRevenue"), "large_currency")),
        ("Net Income (TTM)", _fmt_number(info.get("netIncomeToCommon"), "large_currency")),
        ("Profit Margin", _fmt_number(info.get("profitMargins"), "percent")),
        ("Operating Margin", _fmt_number(info.get("operatingMargins"), "percent")),
        ("ROE", _fmt_number(info.get("returnOnEquity"), "percent")),
        ("ROA", _fmt_number(info.get("returnOnAssets"), "percent")),
        ("Debt to Equity", _fmt_number(info.get("debtToEquity"))),
        ("Current Ratio", _fmt_number(info.get("currentRatio"), "ratio")),
    ]

    row = 3
    headers = ["Field", "Value"]
    for c, h in enumerate(headers, 1):
        ws6.cell(row=row, column=c, value=h)
    _style_header(ws6, row, 2)

    for i, (field, value) in enumerate(profile_fields):
        r = row + 1 + i
        ws6.cell(row=r, column=1, value=field)
        ws6.cell(row=r, column=2, value=value)
        if i % 2 == 0:
            ws6.cell(row=r, column=1).fill = alt_fill
            ws6.cell(row=r, column=2).fill = alt_fill

    # Business summary
    summary_row = row + len(profile_fields) + 2
    ws6.cell(row=summary_row, column=1, value="Business Summary").font = Font(bold=True)
    summary = info.get("longBusinessSummary", "N/A")
    ws6.merge_cells(f"A{summary_row + 1}:D{summary_row + 5}")
    cell = ws6.cell(row=summary_row + 1, column=1, value=summary)
    cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws6.column_dimensions["A"].width = 24
    ws6.column_dimensions["B"].width = 20

    # ── Sheet 9: Price History Chart Data ──
    ws7 = wb.create_sheet("Price History")
    ws7.sheet_properties.tabColor = "5B9BD5"

    hist_raw = data.get("history", {})
    if hist_raw and "Close" in hist_raw:
        ws7["A1"].value = "Historical Price Data (Last 252 Trading Days)"
        ws7["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

        dates = hist_raw.get("Date", [])
        closes = hist_raw.get("Close", [])
        volumes = hist_raw.get("Volume", [])

        # Last ~1 year
        n = min(252, len(closes))
        start_idx = len(closes) - n

        row = 3
        headers = ["Date", "Close", "Volume"]
        for c, h in enumerate(headers, 1):
            ws7.cell(row=row, column=c, value=h)
        _style_header(ws7, row, 3)

        for i in range(n):
            idx = start_idx + i
            r = row + 1 + i
            date_val = dates[idx] if idx < len(dates) else ""
            if isinstance(date_val, str) and "T" in date_val:
                date_val = date_val.split("T")[0]
            ws7.cell(row=r, column=1, value=str(date_val))
            ws7.cell(row=r, column=2, value=round(closes[idx], 2) if idx < len(closes) else 0)
            ws7.cell(row=r, column=3, value=int(volumes[idx]) if idx < len(volumes) and volumes[idx] else 0)

        # Add line chart for close prices
        if n > 10:
            chart = LineChart()
            chart.title = f"{ticker} — Closing Price (1Y)"
            chart.style = 10
            chart.y_axis.title = "Price ($)"
            chart.x_axis.title = "Date"
            chart.width = 30
            chart.height = 15

            data_ref = Reference(ws7, min_col=2, min_row=row,
                                 max_row=row + n)
            chart.add_data(data_ref, titles_from_data=True)
            chart.series[0].graphicalProperties.line.width = 15000

            ws7.add_chart(chart, f"E{row}")

        ws7.column_dimensions["A"].width = 14
        ws7.column_dimensions["B"].width = 12
        ws7.column_dimensions["C"].width = 14
    else:
        ws7["A1"].value = "Price History — No Data Available"

    # ── Sheet 10: Comparable Companies Detail ──
    ws_comps = wb.create_sheet("Comps Detail")
    ws_comps.sheet_properties.tabColor = "4472C4"

    comps_result = next((r for r in results if r.method == "Comps"), None)
    ws_comps["A1"].value = "Comparable Companies Analysis"
    ws_comps["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

    if comps_result and comps_result.fair_value > 0:
        cd = comps_result.details
        row = 3
        ws_comps.cell(row=row, column=1, value="Peers Analyzed")
        ws_comps.cell(row=row, column=2, value=", ".join(cd.get("peers", [])))
        ws_comps.cell(row=row + 1, column=1, value="Current Price")
        ws_comps.cell(row=row + 1, column=2,
                      value=_fmt_number(cd.get("current_price"), "currency"))
        ws_comps.cell(row=row + 2, column=1, value="Comps Fair Value")
        ws_comps.cell(row=row + 2, column=2,
                      value=_fmt_number(comps_result.fair_value, "currency"))
        ws_comps.cell(row=row + 2, column=2).font = Font(bold=True, size=12)

        # Multiples breakdown
        mult_row = row + 4
        mult_headers = ["Multiple", "Target Value", "Peer Median", "Implied Value"]
        for c, h in enumerate(mult_headers, 1):
            ws_comps.cell(row=mult_row, column=c, value=h)
        _style_header(ws_comps, mult_row, len(mult_headers))

        multiples = cd.get("multiples", {})
        for i, (mult_name, md) in enumerate(multiples.items()):
            r = mult_row + 1 + i
            ws_comps.cell(row=r, column=1, value=mult_name)
            ws_comps.cell(row=r, column=2, value=_fmt_number(md.get("target"), "ratio"))
            ws_comps.cell(row=r, column=3, value=_fmt_number(md.get("peer_median"), "ratio"))
            ws_comps.cell(row=r, column=4,
                          value=_fmt_number(md.get("implied_value"), "currency"))
            if i % 2 == 0:
                for c in range(1, 5):
                    ws_comps.cell(row=r, column=c).fill = alt_fill
    else:
        err = comps_result.details.get("error", "N/A") if comps_result else "Not computed"
        ws_comps["A3"].value = f"Insufficient data: {err}"

    for c in range(1, 5):
        ws_comps.column_dimensions[get_column_letter(c)].width = 18

    # ── Sheet 11: Analyst Targets Detail ──
    ws_analyst = wb.create_sheet("Analyst Targets")
    ws_analyst.sheet_properties.tabColor = "ED7D31"

    analyst_result = next((r for r in results if r.method == "Analyst Targets"), None)
    ws_analyst["A1"].value = "Analyst Price Targets"
    ws_analyst["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

    if analyst_result and analyst_result.fair_value > 0:
        ad = analyst_result.details
        row = 3
        analyst_metrics = [
            ("Target Mean", _fmt_number(ad.get("target_mean"), "currency")),
            ("Target Median", _fmt_number(ad.get("target_median"), "currency")),
            ("Target High", _fmt_number(ad.get("target_high"), "currency")),
            ("Target Low", _fmt_number(ad.get("target_low"), "currency")),
            ("Number of Analysts", ad.get("num_analysts", "N/A")),
            ("Confidence", _fmt_number(analyst_result.confidence, "percent")),
        ]
        headers = ["Metric", "Value"]
        for c, h in enumerate(headers, 1):
            ws_analyst.cell(row=row, column=c, value=h)
        _style_header(ws_analyst, row, 2)
        for i, (m, v) in enumerate(analyst_metrics):
            r = row + 1 + i
            ws_analyst.cell(row=r, column=1, value=m)
            ws_analyst.cell(row=r, column=2, value=v)
            if i % 2 == 0:
                ws_analyst.cell(row=r, column=1).fill = alt_fill
                ws_analyst.cell(row=r, column=2).fill = alt_fill

        # Recommendation summary
        rec_row = row + len(analyst_metrics) + 2
        ws_analyst.cell(row=rec_row, column=1,
                        value="Recent Recommendation Breakdown").font = Font(bold=True)
        rec_summary = ad.get("recommendation_summary", {})
        for i, (grade, count) in enumerate(rec_summary.items()):
            r = rec_row + 1 + i
            ws_analyst.cell(row=r, column=1, value=grade.capitalize())
            ws_analyst.cell(row=r, column=2, value=count)
            if grade == "buy":
                ws_analyst.cell(row=r, column=1).font = pos_font
            elif grade == "sell":
                ws_analyst.cell(row=r, column=1).font = neg_font
    else:
        err = analyst_result.details.get("error", "N/A") if analyst_result else "Not computed"
        ws_analyst["A3"].value = f"Insufficient data: {err}"

    ws_analyst.column_dimensions["A"].width = 30
    ws_analyst.column_dimensions["B"].width = 16

    # ── Sheet 12: Historical P/E Valuation ──
    ws_hpe = wb.create_sheet("Historical P-E")
    ws_hpe.sheet_properties.tabColor = "9DC3E6"

    hpe_result = next((r for r in results if r.method == "Historical P/E"), None)
    ws_hpe["A1"].value = "Historical P/E Valuation"
    ws_hpe["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

    if hpe_result and hpe_result.fair_value > 0:
        hd = hpe_result.details
        row = 3
        hpe_metrics = [
            ("Current Price", _fmt_number(hd.get("current_price"), "currency")),
            ("Current P/E", _fmt_number(hd.get("current_pe"), "ratio")),
            ("Historical Avg P/E", _fmt_number(hd.get("historical_avg_pe"), "ratio")),
            ("Trailing EPS", _fmt_number(hd.get("trailing_eps"), "currency")),
            ("Forward EPS", _fmt_number(hd.get("forward_eps"), "currency")),
            ("EPS Used", _fmt_number(hd.get("eps_used"), "currency")),
            ("P/E Data Points", hd.get("pe_data_points", "N/A")),
            ("Fair Value", _fmt_number(hpe_result.fair_value, "currency")),
            ("Confidence", _fmt_number(hpe_result.confidence, "percent")),
        ]
        headers = ["Metric", "Value"]
        for c, h in enumerate(headers, 1):
            ws_hpe.cell(row=row, column=c, value=h)
        _style_header(ws_hpe, row, 2)
        for i, (m, v) in enumerate(hpe_metrics):
            r = row + 1 + i
            ws_hpe.cell(row=r, column=1, value=m)
            ws_hpe.cell(row=r, column=2, value=v)
            if i % 2 == 0:
                ws_hpe.cell(row=r, column=1).fill = alt_fill
                ws_hpe.cell(row=r, column=2).fill = alt_fill
    else:
        err = hpe_result.details.get("error", "N/A") if hpe_result else "Not computed"
        ws_hpe["A3"].value = f"Insufficient data: {err}"

    ws_hpe.column_dimensions["A"].width = 24
    ws_hpe.column_dimensions["B"].width = 16

    # ── Sheet 13: Sentiment Detail ──
    ws_sent = wb.create_sheet("Sentiment Analysis")
    ws_sent.sheet_properties.tabColor = "FF6F61"

    sent_result = next((r for r in results if r.method == "Sentiment"), None)
    ws_sent["A1"].value = "Sentiment Analysis"
    ws_sent["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

    if sent_result and sent_result.fair_value > 0:
        sd_s = sent_result.details
        row = 3
        sent_metrics = [
            ("Current Price", _fmt_number(sd_s.get("current_price"), "currency")),
            ("Headlines Analyzed", sd_s.get("headline_count", 0)),
            ("Composite Sentiment Score", _fmt_number(sd_s.get("sentiment_score"))),
            ("Price Adjustment Factor", _fmt_number(sd_s.get("adjustment_factor"))),
            ("Sentiment Fair Value", _fmt_number(sent_result.fair_value, "currency")),
            ("Confidence", _fmt_number(sent_result.confidence, "percent")),
        ]
        headers = ["Metric", "Value"]
        for c, h in enumerate(headers, 1):
            ws_sent.cell(row=row, column=c, value=h)
        _style_header(ws_sent, row, 2)
        for i, (m, v) in enumerate(sent_metrics):
            r = row + 1 + i
            ws_sent.cell(row=r, column=1, value=m)
            ws_sent.cell(row=r, column=2, value=v)
            if i % 2 == 0:
                ws_sent.cell(row=r, column=1).fill = alt_fill
                ws_sent.cell(row=r, column=2).fill = alt_fill

        # Individual scores
        sc_row = row + len(sent_metrics) + 2
        ws_sent.cell(row=sc_row, column=1,
                     value="Individual Score Components").font = Font(bold=True)
        indiv = sd_s.get("individual_scores", [])
        score_labels = ["News Sentiment", "Profit Margin", "Revenue Growth",
                        "Earnings Growth", "Institutional Holding", "Insider Holding"]
        for i, score in enumerate(indiv):
            r = sc_row + 1 + i
            label = score_labels[i] if i < len(score_labels) else f"Signal {i+1}"
            ws_sent.cell(row=r, column=1, value=label)
            cell = ws_sent.cell(row=r, column=2, value=_fmt_number(score))
            if score > 0:
                cell.font = pos_font
            elif score < 0:
                cell.font = neg_font
    else:
        err = sent_result.details.get("error", "N/A") if sent_result else "Not computed"
        ws_sent["A3"].value = f"Insufficient data: {err}"

    ws_sent.column_dimensions["A"].width = 28
    ws_sent.column_dimensions["B"].width = 16

    # ── Sheet 14: Quality Score ──
    ws_qual = wb.create_sheet("Quality Score")
    ws_qual.sheet_properties.tabColor = "00B050"

    ws_qual["A1"].value = "Company Quality Score"
    ws_qual["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

    if quality_data:
        # Overall grade
        row = 3
        ws_qual.merge_cells(f"A{row}:B{row}")
        grade_cell = ws_qual.cell(row=row, column=1,
                                  value=f"Overall Grade: {quality_data['grade']}  "
                                        f"({quality_data['composite']}/100)")
        grade_cell.font = Font(size=16, bold=True,
                               color=cfg.COLOR_SCHEME["positive"]
                               if quality_data["composite"] >= 65
                               else cfg.COLOR_SCHEME["negative"])

        row = 5
        headers = ["Dimension", "Score (/100)", "Weight"]
        for c, h in enumerate(headers, 1):
            ws_qual.cell(row=row, column=c, value=h)
        _style_header(ws_qual, row, 3)

        scores = quality_data.get("scores", {})
        for i, (dim, weight) in enumerate(cfg.QUALITY_WEIGHTS.items()):
            r = row + 1 + i
            ws_qual.cell(row=r, column=1, value=dim.replace("_", " ").title())
            score_val = scores.get(dim, 0)
            sc = ws_qual.cell(row=r, column=2, value=score_val)
            ws_qual.cell(row=r, column=3, value=_fmt_number(weight, "percent"))
            if score_val >= 70:
                sc.font = pos_font
            elif score_val < 40:
                sc.font = neg_font
            if i % 2 == 0:
                for c in range(1, 4):
                    ws_qual.cell(row=r, column=c).fill = alt_fill

        # Composite row
        r = row + 1 + len(cfg.QUALITY_WEIGHTS)
        ws_qual.cell(row=r + 1, column=1, value="COMPOSITE SCORE").font = Font(bold=True)
        ws_qual.cell(row=r + 1, column=2,
                     value=quality_data["composite"]).font = Font(bold=True, size=14)
    else:
        ws_qual["A3"].value = "Quality score not computed"

    ws_qual.column_dimensions["A"].width = 22
    ws_qual.column_dimensions["B"].width = 16
    ws_qual.column_dimensions["C"].width = 12

    # ── Sheet 15: Scenario Analysis ──
    ws_scen = wb.create_sheet("Scenario Analysis")
    ws_scen.sheet_properties.tabColor = "7030A0"

    ws_scen["A1"].value = "Scenario Analysis (Bear / Base / Bull)"
    ws_scen["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

    if scenario_data and scenario_data.get("base", 0) > 0:
        row = 3
        headers = ["Scenario", "Fair Value", "Upside/Downside", "Weight"]
        for c, h in enumerate(headers, 1):
            ws_scen.cell(row=row, column=c, value=h)
        _style_header(ws_scen, row, 4)

        scenarios = [
            ("Bear", scenario_data["bear"], scenario_data.get("bear_upside", 0),
             cfg.SCENARIO_WEIGHTS["bear"]),
            ("Base", scenario_data["base"], scenario_data.get("base_upside", 0),
             cfg.SCENARIO_WEIGHTS["base"]),
            ("Bull", scenario_data["bull"], scenario_data.get("bull_upside", 0),
             cfg.SCENARIO_WEIGHTS["bull"]),
        ]
        colors = [cfg.COLOR_SCHEME["negative"], cfg.COLOR_SCHEME["neutral"],
                  cfg.COLOR_SCHEME["positive"]]

        for i, (name, fv, upside, weight) in enumerate(scenarios):
            r = row + 1 + i
            ws_scen.cell(row=r, column=1, value=name).font = Font(
                bold=True, color=colors[i])
            ws_scen.cell(row=r, column=2, value=_fmt_number(fv, "currency"))
            up_cell = ws_scen.cell(row=r, column=3, value=_fmt_number(upside, "percent"))
            up_cell.font = pos_font if upside >= 0 else neg_font
            ws_scen.cell(row=r, column=4, value=_fmt_number(weight, "percent"))

        r = row + 5
        ws_scen.cell(row=r, column=1, value="Weighted Scenario Value").font = Font(bold=True)
        ws_scen.cell(row=r, column=2,
                     value=_fmt_number(scenario_data["weighted"], "currency"))
        ws_scen.cell(row=r, column=2).font = Font(bold=True, size=12)
    else:
        ws_scen["A3"].value = "Scenario analysis data unavailable"

    for c in range(1, 5):
        ws_scen.column_dimensions[get_column_letter(c)].width = 20

    # ── Sheet 16: Growth & Margin Sensitivity ──
    ws_gm = wb.create_sheet("Growth-Margin Sensitivity")
    ws_gm.sheet_properties.tabColor = "FFC000"

    ws_gm["A1"].value = "Revenue Growth vs. EBIT Margin Sensitivity"
    ws_gm["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])
    ws_gm["A2"].value = ("Shows implied EV based on varying growth and margin assumptions. "
                         "Uses a simple EV/Revenue multiple approach.")
    ws_gm["A2"].font = Font(size=9, italic=True, color=cfg.COLOR_SCHEME["neutral"])

    dcf_r = next((r for r in results if r.method == "DCF"), None)
    if dcf_r and dcf_r.fair_value > 0:
        d = dcf_r.details
        shares = d.get("shares_outstanding", 1)
        revenue = info.get("totalRevenue", 0) or 0
        wacc = d.get("wacc", 0.10)

        growth_steps = cfg.GROWTH_MARGIN_SENSITIVITY["growth_steps"]
        margin_steps = cfg.GROWTH_MARGIN_SENSITIVITY["margin_steps"]

        row = 4
        ws_gm.cell(row=row, column=1, value="Growth \\ Margin")
        for j, margin in enumerate(margin_steps):
            ws_gm.cell(row=row, column=2 + j, value=f"{margin:.0%}")
        _style_header(ws_gm, row, 1 + len(margin_steps))

        good_fill_gm = PatternFill(start_color=cfg.COLOR_SCHEME["good_bg"],
                                   end_color=cfg.COLOR_SCHEME["good_bg"],
                                   fill_type="solid")
        bad_fill_gm = PatternFill(start_color=cfg.COLOR_SCHEME["bad_bg"],
                                  end_color=cfg.COLOR_SCHEME["bad_bg"],
                                  fill_type="solid")

        for i, growth in enumerate(growth_steps):
            r = row + 1 + i
            ws_gm.cell(row=r, column=1, value=f"{growth:.0%}").font = Font(bold=True)
            for j, margin in enumerate(margin_steps):
                # Project 5Y revenue, apply margin, discount
                if revenue > 0:
                    proj_rev = revenue * ((1 + growth) ** 5)
                    ebit = proj_rev * margin
                    # Simple EV approximation
                    ev_approx = ebit * (1 / wacc) if wacc > 0 else 0
                    cash = info.get("totalCash", 0) or 0
                    debt = info.get("totalDebt", 0) or 0
                    eq_val = ev_approx + cash - debt
                    fv_per_share = eq_val / shares if shares > 0 else 0
                else:
                    fv_per_share = 0

                cell = ws_gm.cell(row=r, column=2 + j,
                                  value=_fmt_number(fv_per_share, "currency"))
                if current_price and fv_per_share > 0:
                    pct = (fv_per_share - current_price) / current_price
                    if pct > 0.15:
                        cell.fill = good_fill_gm
                    elif pct < -0.15:
                        cell.fill = bad_fill_gm
    else:
        ws_gm["A4"].value = "Insufficient DCF data for sensitivity analysis"

    ws_gm.column_dimensions["A"].width = 18
    for j in range(len(cfg.GROWTH_MARGIN_SENSITIVITY.get("margin_steps", []))):
        ws_gm.column_dimensions[get_column_letter(2 + j)].width = 14

    # ── Sheet 17: Income Statement ──
    ws_is = wb.create_sheet("Income Statement")
    ws_is.sheet_properties.tabColor = "5B9BD5"

    ws_is["A1"].value = "Income Statement (Annual)"
    ws_is["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

    is_data = data.get("financials", {}).get("income_stmt", {})
    if is_data:
        is_df = pd.DataFrame(is_data)
        row = 3
        # Column headers = fiscal years
        ws_is.cell(row=row, column=1, value="Line Item")
        for j, col_name in enumerate(is_df.columns):
            col_label = str(col_name)[:10] if len(str(col_name)) > 10 else str(col_name)
            ws_is.cell(row=row, column=2 + j, value=col_label)
        _style_header(ws_is, row, 1 + len(is_df.columns))

        for i, (idx, row_data) in enumerate(is_df.iterrows()):
            r = row + 1 + i
            ws_is.cell(row=r, column=1, value=str(idx))
            for j, val in enumerate(row_data.values):
                try:
                    fval = float(val) if val is not None else None
                    ws_is.cell(row=r, column=2 + j,
                               value=_fmt_number(fval, "large_currency") if fval else "N/A")
                except (TypeError, ValueError):
                    ws_is.cell(row=r, column=2 + j, value=str(val) if val else "N/A")
            if i % 2 == 0:
                for c in range(1, 2 + len(is_df.columns)):
                    ws_is.cell(row=r, column=c).fill = alt_fill

        ws_is.column_dimensions["A"].width = 35
        for j in range(len(is_df.columns)):
            ws_is.column_dimensions[get_column_letter(2 + j)].width = 18
    else:
        ws_is["A3"].value = "Income statement data not available"

    # ── Sheet 18: Balance Sheet ──
    ws_bs = wb.create_sheet("Balance Sheet")
    ws_bs.sheet_properties.tabColor = "00B050"

    ws_bs["A1"].value = "Balance Sheet (Annual)"
    ws_bs["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

    bs_data = data.get("financials", {}).get("balance_sheet", {})
    if bs_data:
        bs_df = pd.DataFrame(bs_data)
        row = 3
        ws_bs.cell(row=row, column=1, value="Line Item")
        for j, col_name in enumerate(bs_df.columns):
            col_label = str(col_name)[:10] if len(str(col_name)) > 10 else str(col_name)
            ws_bs.cell(row=row, column=2 + j, value=col_label)
        _style_header(ws_bs, row, 1 + len(bs_df.columns))

        for i, (idx, row_data) in enumerate(bs_df.iterrows()):
            r = row + 1 + i
            ws_bs.cell(row=r, column=1, value=str(idx))
            for j, val in enumerate(row_data.values):
                try:
                    fval = float(val) if val is not None else None
                    ws_bs.cell(row=r, column=2 + j,
                               value=_fmt_number(fval, "large_currency") if fval else "N/A")
                except (TypeError, ValueError):
                    ws_bs.cell(row=r, column=2 + j, value=str(val) if val else "N/A")
            if i % 2 == 0:
                for c in range(1, 2 + len(bs_df.columns)):
                    ws_bs.cell(row=r, column=c).fill = alt_fill

        ws_bs.column_dimensions["A"].width = 35
        for j in range(len(bs_df.columns)):
            ws_bs.column_dimensions[get_column_letter(2 + j)].width = 18
    else:
        ws_bs["A3"].value = "Balance sheet data not available"

    # ── Sheet 19: Cash Flow Statement ──
    ws_cf = wb.create_sheet("Cash Flow")
    ws_cf.sheet_properties.tabColor = "ED7D31"

    ws_cf["A1"].value = "Cash Flow Statement (Annual)"
    ws_cf["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

    cf_data = data.get("financials", {}).get("cashflow", {})
    if cf_data:
        cf_df = pd.DataFrame(cf_data)
        row = 3
        ws_cf.cell(row=row, column=1, value="Line Item")
        for j, col_name in enumerate(cf_df.columns):
            col_label = str(col_name)[:10] if len(str(col_name)) > 10 else str(col_name)
            ws_cf.cell(row=row, column=2 + j, value=col_label)
        _style_header(ws_cf, row, 1 + len(cf_df.columns))

        for i, (idx, row_data) in enumerate(cf_df.iterrows()):
            r = row + 1 + i
            ws_cf.cell(row=r, column=1, value=str(idx))
            for j, val in enumerate(row_data.values):
                try:
                    fval = float(val) if val is not None else None
                    ws_cf.cell(row=r, column=2 + j,
                               value=_fmt_number(fval, "large_currency") if fval else "N/A")
                except (TypeError, ValueError):
                    ws_cf.cell(row=r, column=2 + j, value=str(val) if val else "N/A")
            if i % 2 == 0:
                for c in range(1, 2 + len(cf_df.columns)):
                    ws_cf.cell(row=r, column=c).fill = alt_fill

        ws_cf.column_dimensions["A"].width = 35
        for j in range(len(cf_df.columns)):
            ws_cf.column_dimensions[get_column_letter(2 + j)].width = 18
    else:
        ws_cf["A3"].value = "Cash flow statement data not available"

    # ── Sheet 20: Peer Comparison Matrix ──
    ws_peer = wb.create_sheet("Peer Comparison")
    ws_peer.sheet_properties.tabColor = "C55A11"

    ws_peer["A1"].value = f"Peer Comparison — {ticker}"
    ws_peer["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

    if comps_result and comps_result.details.get("peers"):
        peer_tickers = comps_result.details["peers"]
        row = 3
        peer_headers = ["Metric", ticker] + peer_tickers
        for c, h in enumerate(peer_headers, 1):
            ws_peer.cell(row=row, column=c, value=h)
        _style_header(ws_peer, row, len(peer_headers))

        compare_metrics = [
            ("Market Cap", "marketCap", "large_currency"),
            ("P/E (Trailing)", "trailingPE", "ratio"),
            ("P/E (Forward)", "forwardPE", "ratio"),
            ("EV/EBITDA", "enterpriseToEbitda", "ratio"),
            ("P/S", "priceToSalesTrailing12Months", "ratio"),
            ("P/B", "priceToBook", "ratio"),
            ("PEG", "pegRatio", "ratio"),
            ("Profit Margin", "profitMargins", "percent"),
            ("Operating Margin", "operatingMargins", "percent"),
            ("ROE", "returnOnEquity", "percent"),
            ("Revenue Growth", "revenueGrowth", "percent"),
            ("Debt/Equity", "debtToEquity", "ratio"),
            ("Dividend Yield", "dividendYield", "percent"),
            ("Beta", "beta", "ratio"),
        ]

        # Cache peer info
        peer_infos = {}
        for pt in peer_tickers:
            try:
                peer_infos[pt] = yf.Ticker(pt).info or {}
            except Exception:
                peer_infos[pt] = {}

        for i, (label, key, fmt) in enumerate(compare_metrics):
            r = row + 1 + i
            ws_peer.cell(row=r, column=1, value=label)
            # Target value
            target_val = info.get(key)
            ws_peer.cell(row=r, column=2,
                         value=_fmt_number(target_val, fmt) if target_val else "N/A")
            # Peer values
            for j, pt in enumerate(peer_tickers):
                peer_val = peer_infos.get(pt, {}).get(key)
                ws_peer.cell(row=r, column=3 + j,
                             value=_fmt_number(peer_val, fmt) if peer_val else "N/A")
            if i % 2 == 0:
                for c in range(1, len(peer_headers) + 1):
                    ws_peer.cell(row=r, column=c).fill = alt_fill

        ws_peer.column_dimensions["A"].width = 20
        for c in range(2, len(peer_headers) + 1):
            ws_peer.column_dimensions[get_column_letter(c)].width = 14
    else:
        ws_peer["A3"].value = "No peer data available for comparison"

    # ── Sheet 21: Risk Assessment ──
    ws_risk = wb.create_sheet("Risk Assessment")
    ws_risk.sheet_properties.tabColor = "FF0000"

    ws_risk["A1"].value = "Risk Assessment"
    ws_risk["A1"].font = Font(size=14, bold=True, color=cfg.COLOR_SCHEME["header_bg"])

    row = 3
    risk_headers = ["Risk Factor", "Level", "Details"]
    for c, h in enumerate(risk_headers, 1):
        ws_risk.cell(row=row, column=c, value=h)
    _style_header(ws_risk, row, 3)

    risks = []
    # Valuation risk
    if current_price and composite["composite_fair_value"] > 0:
        pct_diff = (composite["composite_fair_value"] - current_price) / current_price
        if pct_diff < -0.20:
            risks.append(("Valuation Risk", "High",
                          f"Stock appears {abs(pct_diff):.0%} overvalued vs fair value"))
        elif pct_diff < -0.05:
            risks.append(("Valuation Risk", "Medium",
                          f"Stock may be {abs(pct_diff):.0%} overvalued"))
        else:
            risks.append(("Valuation Risk", "Low",
                          "Stock appears fairly valued or undervalued"))

    # Volatility risk
    beta_val = info.get("beta", 1.0) or 1.0
    if beta_val > 1.5:
        risks.append(("Volatility Risk", "High",
                      f"Beta of {beta_val:.2f} indicates high market sensitivity"))
    elif beta_val > 1.0:
        risks.append(("Volatility Risk", "Medium",
                      f"Beta of {beta_val:.2f}, slightly above market"))
    else:
        risks.append(("Volatility Risk", "Low",
                      f"Beta of {beta_val:.2f}, less volatile than market"))

    # Leverage risk
    de = info.get("debtToEquity", 0) or 0
    if de > 10:
        de = de / 100.0
    if de > 2.0:
        risks.append(("Leverage Risk", "High",
                      f"Debt/Equity of {de:.2f} is significantly elevated"))
    elif de > 1.0:
        risks.append(("Leverage Risk", "Medium",
                      f"Debt/Equity of {de:.2f}, moderate leverage"))
    else:
        risks.append(("Leverage Risk", "Low",
                      f"Debt/Equity of {de:.2f}, conservative balance sheet"))

    # Liquidity risk
    cr = info.get("currentRatio", 0) or 0
    if cr < 1.0:
        risks.append(("Liquidity Risk", "High",
                      f"Current ratio of {cr:.2f} below 1.0"))
    elif cr < 1.5:
        risks.append(("Liquidity Risk", "Medium",
                      f"Current ratio of {cr:.2f}, adequate but watch"))
    else:
        risks.append(("Liquidity Risk", "Low",
                      f"Current ratio of {cr:.2f}, strong liquidity"))

    # Concentration risk
    inst_pct = info.get("heldPercentInstitutions", 0) or 0
    if inst_pct > 0.90:
        risks.append(("Ownership Concentration", "Medium",
                      f"{inst_pct:.0%} institutional ownership — crowded"))
    else:
        risks.append(("Ownership Concentration", "Low",
                      f"{inst_pct:.0%} institutional ownership"))

    # Growth risk
    rev_growth = info.get("revenueGrowth", 0) or 0
    if rev_growth < 0:
        risks.append(("Growth Risk", "High",
                      f"Revenue declining at {rev_growth:.1%}"))
    elif rev_growth < 0.05:
        risks.append(("Growth Risk", "Medium",
                      f"Slow revenue growth of {rev_growth:.1%}"))
    else:
        risks.append(("Growth Risk", "Low",
                      f"Healthy revenue growth of {rev_growth:.1%}"))

    high_fill = PatternFill(start_color=cfg.COLOR_SCHEME["bad_bg"],
                            end_color=cfg.COLOR_SCHEME["bad_bg"], fill_type="solid")
    med_fill = PatternFill(start_color=cfg.COLOR_SCHEME["warn_bg"],
                           end_color=cfg.COLOR_SCHEME["warn_bg"], fill_type="solid")
    low_fill = PatternFill(start_color=cfg.COLOR_SCHEME["good_bg"],
                           end_color=cfg.COLOR_SCHEME["good_bg"], fill_type="solid")

    for i, (factor, level, detail) in enumerate(risks):
        r = row + 1 + i
        ws_risk.cell(row=r, column=1, value=factor)
        level_cell = ws_risk.cell(row=r, column=2, value=level)
        ws_risk.cell(row=r, column=3, value=detail)
        fill_map = {"High": high_fill, "Medium": med_fill, "Low": low_fill}
        for c in range(1, 4):
            ws_risk.cell(row=r, column=c).fill = fill_map.get(level, alt_fill)

    ws_risk.column_dimensions["A"].width = 24
    ws_risk.column_dimensions["B"].width = 12
    ws_risk.column_dimensions["C"].width = 50

    # Save
    wb.save(filepath)
    return filepath


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 7 — Console Report
# ═══════════════════════════════════════════════════════════════════════════════

def print_console_summary(ticker: str, data: dict, results: list,
                          composite: dict, quality_data: dict = None,
                          scenario_data: dict = None):
    """Print a concise summary to the console."""
    info = data.get("info", {})
    current_price = info.get("currentPrice") or info.get("regularMarketPrice", 0)
    fair_value = composite["composite_fair_value"]

    print("\n" + "=" * 70)
    print(f"  {ticker} — {info.get('shortName', ticker)}")
    print(f"  {info.get('sector', '')} | {info.get('industry', '')}")
    print("=" * 70)
    print(f"  Current Price:          ${current_price:>10,.2f}")
    print(f"  Composite Fair Value:   ${fair_value:>10,.2f}")

    if current_price > 0:
        pct = (fair_value - current_price) / current_price
        direction = "Upside" if pct >= 0 else "Downside"
        print(f"  {direction}:                 {pct:>10.1%}")

    # Margin of safety
    mos = cfg.MARGIN_OF_SAFETY
    buy_below = fair_value * (1 - mos)
    print(f"  Buy Below (w/ {mos:.0%} MoS): ${buy_below:>10,.2f}")

    print("-" * 70)
    print("  Valuation Method Breakdown:")
    for method, bd in composite["breakdown"].items():
        fv = bd["fair_value"]
        conf = bd["confidence"]
        note = bd.get("note", "")
        wt = bd.get("weight", 0)
        if fv > 0:
            print(f"    {method:<20s}  ${fv:>10,.2f}  (conf: {conf:.0%}, wt: {wt:.0%})")
        else:
            print(f"    {method:<20s}  {'N/A':>11s}  — {note}")

    # Quality Score
    if quality_data:
        print("-" * 70)
        print(f"  Quality Score:  {quality_data['composite']}/100  "
              f"(Grade: {quality_data['grade']})")
        scores = quality_data.get("scores", {})
        for dim, score in scores.items():
            print(f"    {dim.replace('_', ' ').title():<24s}  {score:>3d}/100")

    # Scenario Analysis
    if scenario_data and scenario_data.get("base", 0) > 0:
        print("-" * 70)
        print("  Scenario Analysis:")
        print(f"    Bear Case:   ${scenario_data['bear']:>10,.2f}  "
              f"({scenario_data.get('bear_upside', 0):>+.1%})")
        print(f"    Base Case:   ${scenario_data['base']:>10,.2f}  "
              f"({scenario_data.get('base_upside', 0):>+.1%})")
        print(f"    Bull Case:   ${scenario_data['bull']:>10,.2f}  "
              f"({scenario_data.get('bull_upside', 0):>+.1%})")
        print(f"    Weighted:    ${scenario_data['weighted']:>10,.2f}")

    print("-" * 70)
    if current_price > 0:
        pct = (fair_value - current_price) / current_price
        if pct > 0.20:
            print("  >> STRONG BUY — Significantly undervalued")
        elif pct > 0.10:
            print("  >> BUY — Undervalued with margin of safety")
        elif pct > -0.10:
            print("  >> HOLD — Fairly valued")
        elif pct > -0.20:
            print("  >> SELL — Appears overvalued")
        else:
            print("  >> STRONG SELL — Significantly overvalued")
    print("=" * 70 + "\n")


# ═══════════════════════════════════════════════════════════════════════════════
# SECTION 8 — Main Orchestrator
# ═══════════════════════════════════════════════════════════════════════════════

def analyze_stock(ticker: str) -> str:
    """Run full analysis for a single ticker. Returns path to Excel report."""
    ticker = ticker.upper().strip()
    print(f"\n{'━' * 60}")
    print(f"  Analyzing {ticker}...")
    print(f"{'━' * 60}")

    # Step 1: Fetch data
    print("\n[ 1/11] Fetching market data...")
    data = fetch_yfinance_data(ticker)

    if not data.get("info"):
        print(f"  ERROR: Could not retrieve data for {ticker}. Skipping.")
        return ""

    # Step 2-8: Run valuation models
    print("[ 2/11] Running DCF valuation...")
    dcf = dcf_valuation(data)

    print("[ 3/11] Running comparable companies analysis...")
    comps = comps_valuation(data, ticker)

    print("[ 4/11] Gathering analyst price targets...")
    analyst = analyst_valuation(data)

    print("[ 5/11] Running technical analysis...")
    tech = technical_valuation(data)

    print("[ 6/11] Analyzing sentiment...")
    sent = sentiment_valuation(data, ticker)

    print("[ 7/11] Analyzing seasonal patterns...")
    seas = seasonal_valuation(data)

    print("[ 8/11] Running Historical P/E valuation...")
    hist_pe = historical_pe_valuation(data)

    results = [dcf, comps, analyst, tech, sent, seas, hist_pe]

    # Step 9: Composite (only uses DCF, Comps, Historical P/E for fair value)
    composite = composite_valuation(results, cfg.VALUATION_WEIGHTS)

    # Quality & Scenario
    print("[ 9/11] Computing quality score...")
    quality_data = quality_score_analysis(data)

    print("[10/11] Running scenario analysis...")
    scenario_data = scenario_analysis(data, results)

    # Macro
    macro = fetch_macro_environment()

    # Console output
    print_console_summary(ticker, data, results, composite,
                          quality_data, scenario_data)

    # Excel report
    print("[11/11] Generating Excel report (20 sheets)...")
    filepath = generate_excel_report(ticker, data, results, composite, macro,
                                     quality_data, scenario_data)
    print(f"  Report saved to: {filepath}")

    return filepath


def main():
    parser = argparse.ArgumentParser(
        description="Stock Valuation Analyzer — multi-method fair value estimation"
    )
    parser.add_argument("tickers", nargs="*",
                        help="One or more stock tickers to analyze (e.g. AAPL MSFT)")
    args = parser.parse_args()

    tickers = args.tickers
    if not tickers:
        raw = input("Enter ticker(s) separated by spaces: ").strip()
        tickers = raw.upper().split()

    if not tickers:
        print("No tickers provided. Exiting.")
        sys.exit(1)

    print(f"\nStock Valuation Analyzer")
    print(f"Tickers: {', '.join(tickers)}")
    print(f"Date: {datetime.datetime.now():%Y-%m-%d %H:%M}\n")

    reports = []
    for t in tickers:
        path = analyze_stock(t)
        if path:
            reports.append(path)

    if reports:
        print("\n" + "=" * 60)
        print("  Generated Reports:")
        for rp in reports:
            print(f"    {rp}")
        print("=" * 60)
    else:
        print("\nNo reports were generated.")


if __name__ == "__main__":
    main()
